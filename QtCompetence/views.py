# is_department is_position 第一万条记录处理空数据手动创建
from django.shortcuts import render,redirect,reverse
from django.contrib.auth import login,logout
from .forms import *
from .models import *
from itertools import chain
# 详情
from django.views.generic.detail import DetailView
# 增删改
from django.views.generic.edit import FormView, CreateView, UpdateView, DeleteView
from django.views.generic import ListView
from django.urls import reverse_lazy
#from django.utils import timezone
from django.http import HttpResponse
import json,datetime,time
from django.db.models import Q
from django.db import connections
# from suds.client import Client  
# url = 'http://tjws001.auto.contiwan.com/windowsauthenticate/service.asmx?wsdl'  
# client = Client(url)  
from apscheduler.schedulers.background import BackgroundScheduler
from django_apscheduler.jobstores import DjangoJobStore, register_events, register_job
# filter
from .filter import audit_info_filter,finding_info_filter,action_info_filter
from django.core.paginator import Paginator
#db
from django.db import connection
#excel export
from openpyxl import Workbook
from io import BytesIO
from django.utils.http import urlquote
#permission
from django.contrib.auth.decorators import permission_required
from django.contrib.auth.mixins import PermissionRequiredMixin
#from django.contrib.auth import authenticate, login
#用户登录
class login_user(FormView):
    template_name = 'QtCompetence/login.html'
    form_class = user_login_form
    success_url = reverse_lazy('QtCompetence:home')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        real_name=self.request.session.get('real_name',None)
        is_manager=self.request.session.get('is_manager',None)
        user_pk = self.request.session.get('user_pk', None)
        is_staff = self.request.session.get('is_staff',None)
        is_departmentManager = self.request.session.get('is_departmentManager',None)
        is_plantManager = self.request.session.get('is_plantManager',None)
        is_group_leader = self.request.session.get('is_group_leader',None)
        context['real_name'] = real_name
        context['is_staff']=is_staff
        context['user_pk']=user_pk
        context['is_manager']=is_manager
        context['is_departmentManager']=is_departmentManager
        context['is_plantManager']=is_plantManager
        context['is_group_leader']=is_group_leader
        return context
    #判断登录信息，校验全部在forms.py中完成
    def post(self,request,*args,**kwargs):
        if request.method == "POST":
            loginForm = user_login_form(request.POST)
            if loginForm.is_valid():
                #login(request, authenticate(self.request, username=self.request.POST['user_name'], password=self.request.POST['user_pwd']),'QtCompetenceBackend')
                user_name = loginForm.cleaned_data['user_name']
                user_pwd = loginForm.cleaned_data['user_pwd']
                real_name = is_user.objects.filter(user_name = user_name)[0].employ_name
                user_pk = is_user.objects.filter(user_name = user_name)[0].id
                is_PM=int(is_plantManager(user_pk))
                is_DPMR=int(is_departmentManager(user_pk))
                is_MR=int(is_manager(user_pk))
                is_SF=int(is_staff(user_pk))
                is_GL=int(is_group_leader(user_pk))
                #如果用户名密码均不为空则登录home主页，并把其写入session
                self.request.session['user_name']=user_name
                self.request.session['user_pk']=user_pk
                self.request.session['real_name']=real_name
                self.request.session['is_staff']=is_SF
                self.request.session['is_manager']=is_MR
                self.request.session['is_departmentManager']=is_DPMR
                self.request.session['is_plantManager']=is_PM
                self.request.session['is_group_leader']=is_GL
                return redirect('QtCompetence:home')
            else:
                return self.form_invalid(loginForm)

#Intelligent Quality 主页
class home(FormView):
    template_name = 'QtCompetence/home.html'
    form_class = user_login_form
    success_url = reverse_lazy('QtCompetence:home')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        real_name=self.request.session.get('real_name',None)
        is_manager=self.request.session.get('is_manager',None)
        user_pk = self.request.session.get('user_pk', None)
        is_staff = self.request.session.get('is_staff',None)
        is_departmentManager = self.request.session.get('is_departmentManager',None)
        is_plantManager = self.request.session.get('is_plantManager',None)
        is_group_leader = self.request.session.get('is_group_leader',None)
        context['real_name'] = real_name
        context['is_staff']=is_staff
        context['user_pk']=user_pk
        context['is_manager']=is_manager
        context['is_departmentManager']=is_departmentManager
        context['is_plantManager']=is_plantManager
        context['is_group_leader']=is_group_leader
        return context
    
    def post(self,request,*args,**kwargs):
        if request.method == "POST":
            loginForm = user_login_form(request.POST)
            if loginForm.is_valid():  # edu-experience
                user_name = loginForm.cleaned_data['user_name']
                user_pwd = loginForm.cleaned_data['user_pwd']
                real_name = is_user.objects.filter(user_name = user_name)[0].employ_name
                user_pk = is_user.objects.filter(user_name = user_name)[0].id
                is_PM=int(is_plantManager(user_pk))
                is_DPMR=int(is_departmentManager(user_pk))
                is_MR=int(is_manager(user_pk))
                is_SF=int(is_staff(user_pk))
                is_GL=int(is_staff(user_pk))
                #如果用户名密码均不为空则登录home主页，并把其写入session
                self.request.session['user_name']=user_name
                self.request.session['user_pk']=user_pk
                self.request.session['real_name']=real_name
                self.request.session['is_staff']=is_SF
                self.request.session['is_manager']=is_MR
                self.request.session['is_departmentManager']=is_DPMR
                self.request.session['is_plantManager']=is_PM
                self.request.session['is_group_leader']=is_GL
                return redirect('QtCompetence:home')
            else:
                return self.form_invalid(loginForm)
#分为四重身份，第一种为普通员工
#is_user中的id与r_level中的level_id为一对一关系，二者永远相等可以互相引用
def is_staff(user_id):
    if not is_plantManager(user_id) and not is_departmentManager(user_id) and not is_manager(user_id) and not is_group_leader(user_id):
        return True
    else:
        return False
#部门主管
#select USER_NAME,EMAIL_ID,INUSE,LEVEL_1_R,LEVEL_2_R,LEVEL_3_R,LEVEL_4_R from ir_dimUser WHERE INUSE='Y' AND LEVEL_1_R in (select ID from ir_dimUser WHERE INUSE='Y' AND LEVEL_1_R=19195 AND LEVEL_4_R=19195) and LEVEL_2_R is NULL and LEVEL_3_R is NULL and LEVEL_4_R=19195;

#group leader 51到54 position_id 目前xu，Darren是特殊情况需改正20200812
def is_group_leader(id):
    if (is_user.objects.filter(id=id)[0].position_id >= 51 and is_user.objects.filter(id=id)[0].position_id <= 54) or is_user.objects.filter(id=id)[0].position_id==144:
        return True
    else:
        return False

def is_manager(id):
    user_id= is_user.objects.filter(user_id=19195)[0].id
    qtDP=r_level.objects.filter(Q(L1_id=user_id) & Q(L4_id=user_id))
    L1_R=r_level.objects.filter(user_id=id)[0].L1_id
    for i in qtDP:
        if L1_R == i.user_id:
            return True
            break
        else:
            continue
    return False

#select EMAIL_ID,USER_NAME,JOB_CODE_ID,LEVEL_1_R,LEVEL_2_R,LEVEL_3_R,LEVEL_4_R from ir_dimUser WHERE INUSE='Y' AND LEVEL_1_R=19195 AND LEVEL_4_R=19195;
#部门经理DP
def is_departmentManager(id):
    user_id= is_user.objects.filter(user_id=19195)[0].id
    qtDP=r_level.objects.filter(Q(L1_id=user_id) & Q(L4_id=user_id))
    for i in qtDP:
        if i.user_id == id:
            return True
            break
        else:
            continue
    return False
#19195
#厂长
def is_plantManager(id):
    user_id=is_user.objects.filter(user_id=19195)[0].id
    if user_id==id:
        return True
    else:
        return False

#@permission_required('QtCompetence.add_is_competition_score')
def competence_judge(request):
#查看经理是否已经评价他的自评，若他的最新评价未被经理评价则可一直更改，一直到经理评价后则无法修改，只可以重新评价。-judgelimit
    user_pk = request.session.get('user_pk',None)
    user_name = request.session.get('user_name',None)
    real_name = request.session.get('real_name',None)
    is_staff = request.session.get('is_staff',None)
    is_manager = request.session.get('is_manager',None)
    is_departmentManager = request.session.get('is_departmentManager',None)
    is_plantManager = request.session.get('is_plantManager',None)
    is_group_leader = request.session.get('is_group_leader',None)
    feedbackInfo=[]
    if user_pk:
        # year=datetime.datetime.now().year
        # month=datetime.datetime.now().month
        if not is_competition_score.objects.filter(user_id=user_pk).exists():
            if request.is_ajax():
                #user gap
                #筛选出某人岗位对应的评价项
                positionID=is_user.objects.filter(id=user_pk)[0].position_id
                qtIDRules=r_rules_position.objects.filter(position_id=positionID)
                rule_ids=[]
                for i in qtIDRules:
                    if is_competition_rule.objects.filter(competence_id=i.rules_id).exists():
                        rule_ids.append(i.rules_id)
                competence_rules = is_competition_rule.objects.filter(competence_id__in=rule_ids).order_by('competence_id')
                score_list = request.POST.get('score_list',None)
                competence_id_list = ','.join([str(q.competence_id) for q in competence_rules])
                #由前端传递而来的数据(原来是字符串数组)转换为列表
                if score_list is not None:
                    score_list = score_list.split(',')
                if competence_id_list is not None:
                    competence_id_list = competence_id_list.split(',')
                qtScores=r_rules_position.objects.filter(position_id=is_user.objects.filter(id=user_pk)[0].position_id)
                if qtScores.count() >= len(competence_id_list):
                    for i in range(len(competence_id_list)):
                        standard_score = r_rules_position.objects.filter(Q(position_id=is_user.objects.filter(id=user_pk)[0].position_id)& Q(rules_id=competence_id_list[i]))[0].standard_score
                        # print(score_list[i])
                        # print(standard_score)
                        gap =int(score_list[i]) - int(standard_score)
                        is_competition_score.objects.create(self_assessment=score_list[i],competence_id=competence_id_list[i],self_gap=gap,user_id=user_pk)
                    return HttpResponse(json.dumps({"info": "评分提交成功",
                                                    "display": "0"}),
                                                    content_type="application/json")
                else:
                    nunms=len(competence_id_list)-qtScores.count()
                    return HttpResponse(json.dumps({"display": "0",
                                                "info": "该岗位针对"+nunms+"条评价规则未设置标准分数,请联系管理员增加标准分数",}),
                                                content_type="application/json")
            else:
                #筛选出某人岗位对应的评价项
                positionID=is_user.objects.filter(id=user_pk)[0].position_id
                qtIDRules=r_rules_position.objects.filter(position_id=positionID).order_by("rules_id")
                qtScores=r_rules_position.objects.filter(position_id=is_user.objects.filter(id=user_pk)[0].position_id)
                rule_ids=[]
                for i in qtIDRules:
                    if is_competition_rule.objects.filter(competence_id=i.rules_id).exists():
                        rule_ids.append(i.rules_id)
                competence_rules = is_competition_rule.objects.filter(competence_id__in=rule_ids).order_by('competence_id')
                #前端样式id对应competenceID所以rule_id也得返回去
                num_rules = competence_rules.count()
                contentForm = judgeForm()
                #均为设置标准分数
                #部分未设置标准分数
                #设置逻辑控制group leader
                if is_group_leader == 1:
                    is_staff = 1
                if qtScores.exists():
                    feedbackInfo={"is_staff":is_staff,"is_manager":is_manager,"is_departmentManager":is_departmentManager,"is_group_leader":is_group_leader,
                                "is_plantManager":is_plantManager,"num_rules":num_rules,'competence_id':rule_ids,"display": "1",
                            "info": ""}
                else:
                    feedbackInfo={"is_staff":is_staff,"is_manager":is_manager,"is_departmentManager":is_departmentManager,"is_group_leader":is_group_leader,
                            "is_plantManager":is_plantManager,"num_rules":num_rules,'competence_id':rule_ids,"display": "0",
                            "info": "该岗位针对所有评价规则未设置标准分数,请联系管理员增加标准分数"}
                return render(request,'QtCompetence/competence_judge.html',{'contentForm':contentForm,
                                                                            'competence_rules':competence_rules,
                                                                            "feedbackInfo": json.dumps(feedbackInfo),
                                                                            "real_name":real_name,
                                                                            "user_pk":user_pk,
                                                                            "is_staff":is_staff,
                                                                            "is_manager":is_manager,
                                                                            "is_departmentManager":is_departmentManager,
                                                                            "is_plantManager":is_plantManager,
                                                                            "is_group_leader":is_group_leader})
        else:
            #如果有记录先查看他的最新评价是否有经理评价分数
            qtScores=is_competition_score.objects.filter(user_id=user_pk)
            #获取某个人所有评价日期的不重复值
            #all_judge_create_times = qtScores.values('create_time').distinct()
            #qtAllJudgeCreateTimes=(qtScores.extra(select={"create_time": "DATE_FORMAT(create_time, '%%Y-%%m-%%d %%H:%%i')"}).values('create_time')).distinct().order_by('-create_time')
            #有一条记录的人
            max_scores_create_time=(qtScores.order_by('-create_time')[:1])[0].create_time
            min_scores_create_time=max_scores_create_time+datetime.timedelta(seconds=-4)
            #切分为两部分，一部分为只可以查看的editScores，另一部分只可以查看readScores
            if qtScores.filter(~(Q(create_time__gte=(min_scores_create_time)) &Q(create_time__lte=(max_scores_create_time)))).exists():
                qtReadScores=qtScores.filter(~(Q(create_time__gte=(min_scores_create_time)) &Q(create_time__lte=(max_scores_create_time)))).order_by('-create_time')
            else:
                qtReadScores=""
            if qtScores.filter(Q(create_time__gte=(min_scores_create_time)) &Q(create_time__lte=(max_scores_create_time))).exists():
                qtEditScores=qtScores.filter(Q(create_time__gte=(min_scores_create_time)) &Q(create_time__lte=(max_scores_create_time)))
                if qtEditScores[0].manager_assessment is not None:
                    if request.is_ajax():
                        #user gap
                        #筛选出某人岗位对应的评价项
                        positionID=is_user.objects.filter(id=user_pk)[0].position_id
                        qtIDRules=r_rules_position.objects.filter(position_id=positionID)
                        rule_ids=[]
                        for i in qtIDRules:
                            if is_competition_rule.objects.filter(competence_id=i.rules_id).exists():
                                rule_ids.append(i.rules_id)
                        competence_rules = is_competition_rule.objects.filter(competence_id__in=rule_ids).order_by('competence_id')
                        score_list = request.POST.get('score_list',None)
                        competence_id_list = ','.join([str(q.competence_id) for q in competence_rules])
                        #由前端传递而来的数据(原来是字符串数组)转换为列表
                        if score_list is not None:
                            score_list = score_list.split(',')
                        if competence_id_list is not None:
                            competence_id_list = competence_id_list.split(',')
                        qtScores=r_rules_position.objects.filter(position_id=is_user.objects.filter(id=user_pk)[0].position_id)
                        if qtScores.count() >= len(competence_id_list):
                            for i in range(len(competence_id_list)):
                                standard_score = r_rules_position.objects.filter(Q(position_id=is_user.objects.filter(id=user_pk)[0].position_id)& Q(rules_id=competence_id_list[i]))[0].standard_score
                                # print(score_list[i])
                                # print(standard_score)
                                gap =int(score_list[i]) - int(standard_score)
                                is_competition_score.objects.create(self_assessment=score_list[i],competence_id=competence_id_list[i],self_gap=gap,user_id=user_pk)
                            return HttpResponse(json.dumps({"info": "评分提交成功",
                                                            "display": "0"}),
                                                            content_type="application/json")
                        else:
                            nunms=len(competence_id_list)-qtScores.count()
                            return HttpResponse(json.dumps({"display": "0",
                                                        "info": "该岗位针对"+nunms+"条评价规则未设置标准分数,请联系管理员增加标准分数",}),
                                                        content_type="application/json")
                    else:
                        #筛选出某人岗位对应的评价项
                        positionID=is_user.objects.filter(id=user_pk)[0].position_id
                        qtIDRules=r_rules_position.objects.filter(position_id=positionID).order_by("rules_id")
                        qtScores=r_rules_position.objects.filter(position_id=is_user.objects.filter(id=user_pk)[0].position_id)
                        rule_ids=[]
                        for i in qtIDRules:
                            if is_competition_rule.objects.filter(competence_id=i.rules_id).exists():
                                rule_ids.append(i.rules_id)
                        competence_rules = is_competition_rule.objects.filter(competence_id__in=rule_ids).order_by('competence_id')
                        #前端样式id对应competenceID所以rule_id也得返回去
                        num_rules = competence_rules.count()
                        contentForm = judgeForm()
                        #均为设置标准分数
                        #部分未设置标准分数
                        #设置逻辑控制group leader
                        if is_group_leader == 1:
                            is_staff = 1
                        if qtScores.exists():
                            feedbackInfo={"is_staff":is_staff,"is_manager":is_manager,"is_departmentManager":is_departmentManager,"is_group_leader":is_group_leader,
                                        "is_plantManager":is_plantManager,"num_rules":num_rules,'competence_id':rule_ids,"display": "1",
                                    "info": ""}
                        else:
                            feedbackInfo={"is_staff":is_staff,"is_manager":is_manager,"is_departmentManager":is_departmentManager,"is_group_leader":is_group_leader,
                                    "is_plantManager":is_plantManager,"num_rules":num_rules,'competence_id':rule_ids,"display": "0",
                                    "info": "该岗位针对所有评价规则未设置标准分数,请联系管理员增加标准分数"}
                        return render(request,'QtCompetence/competence_judge.html',{'contentForm':contentForm,
                                                                                    'competence_rules':competence_rules,
                                                                                    "feedbackInfo": json.dumps(feedbackInfo),
                                                                                    "real_name":real_name,
                                                                                    "user_pk":user_pk,
                                                                                    "is_staff":is_staff,
                                                                                    "is_manager":is_manager,
                                                                                    "is_departmentManager":is_departmentManager,
                                                                                    "is_plantManager":is_plantManager,
                                                                                    "is_group_leader":is_group_leader})
                else:
                    #直接返回可以任意修改
                    return render(request,'QtCompetence/competence_info_list.html',{"qtEditScores":qtEditScores,
                                                                    "qtReadScores":qtReadScores,
                                                                    "real_name":real_name,
                                                                    "user_pk":user_pk,
                                                                    "is_staff":is_staff,
                                                                    "is_manager":is_manager,
                                                                    "is_departmentManager":is_departmentManager,
                                                                    "is_plantManager":is_plantManager,
                                                                    "is_group_leader":is_group_leader})
            else:
                return render(request,'QtCompetence/competence_info_list.html',{"qtEditScores":"",
                                                                    "qtReadScores":qtReadScores,
                                                                    "real_name":real_name,
                                                                    "user_pk":user_pk,
                                                                    "is_staff":is_staff,
                                                                    "is_manager":is_manager,
                                                                    "is_departmentManager":is_departmentManager,
                                                                    "is_plantManager":is_plantManager,
                                                                    "is_group_leader":is_group_leader})

    else:
        return redirect('QtCompetence:login_user')

#is_competition_score update
class competence_info_update(UpdateView):
    template_name = 'QtCompetence/update_competence.html'
    model = is_competition_score
    success_url = reverse_lazy('QtCompetence:competence_judge')
    fields = ['self_assessment',]
    #form_class=auditCreateForm
    queryset=is_competition_score.objects.all().order_by('assessment_id')

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        real_name = self.request.session.get('real_name', None)
        user_pk = self.request.session.get('user_pk', None)
        is_staff = self.request.session.get('is_staff',None)
        is_manager = self.request.session.get('is_manager',None)
        is_departmentManager = self.request.session.get('is_departmentManager',None)
        is_plantManager = self.request.session.get('is_plantManager',None)
        is_group_leader = self.request.session.get('is_group_leader',None)
        context['real_name'] = real_name
        context['is_staff']=is_staff
        context['user_pk']=user_pk
        context['is_manager']=is_manager
        context['is_departmentManager']=is_department
        context['is_plantManager']=is_plantManager
        context['is_group_leader']=is_group_leader
        competence_id=is_competition_score.objects.filter(assessment_id=self.kwargs['pk'])[0].competence_id
        context['competence_rules']=is_competition_rule.objects.filter(competence_id=competence_id)
        return context

    def post(self, request, *args, **kwargs):
        user_pk = self.request.session.get('user_pk', None)
        if user_pk:
            competence_id=is_competition_score.objects.filter(assessment_id=self.kwargs['pk'])[0].competence_id
            self_assessment = request.POST.get("self_assessment", None)
            #找到岗位对应的standard score相减更新gap值
            standard_score=r_rules_position.objects.filter(Q(position_id=is_user.objects.filter(id=user_pk)[0].position_id) & Q(rules_id=competence_id))[0].standard_score
            is_competition_score.objects.filter(assessment_id=self.kwargs['pk']).update(competence_id=competence_id,self_assessment=int(self_assessment),self_gap=int(self_assessment)-standard_score)
            return redirect('QtCompetence:competence_judge')
        else:
            return redirect('QtCompetence:login_user')


#经理评价手下员工 所有评价员工都可以有任意多的记录
def competence_manager_judge(request,user_id):
    user_pk = request.session.get('user_pk',None)
    user_name = request.session.get('user_name',None)
    real_name = request.session.get('real_name',None)
    is_staff = request.session.get('is_staff',None)
    is_manager = request.session.get('is_manager',None)
    is_departmentManager = request.session.get('is_departmentManager',None)
    is_plantManager = request.session.get('is_plantManager',None)
    is_group_leader = request.session.get('is_group_leader',None)
    if user_pk: 
        if request.is_ajax():
            #第一条记录
            #评价最后一次自评结果
            #能进入此处评价则可认为必定已经进行了自评则qtScore不为空
            competence_rules = is_competition_rule.objects.filter(position=is_user.objects.filter(id=user_id)[0].position_id).order_by('competence_id')
            score_list = request.POST.get('score_list',None)
            competence_id_list = ','.join([str(q.competence_id) for q in competence_rules])
            #由前端传递而来的数据(原来是字符串数组)转换为列表
            if score_list is not None:
                score_list = score_list.split(',')
            if competence_id_list is not None:
                competence_id_list = competence_id_list.split(',')
            qtScores=is_competition_score.objects.filter(user_id=user_id)
            max_scores_create_time=(qtScores.order_by('-create_time')[:1])[0].create_time
            min_scores_create_time=max_scores_create_time+datetime.timedelta(seconds=-4)
            #切分为两部分，一部分为只可以查看的editScores，另一部分只可以查看readScores
            if qtScores.filter(Q(create_time__gte=(min_scores_create_time)) &Q(create_time__lte=(max_scores_create_time))).exists():
                qtEditScores=qtScores.filter(Q(create_time__gte=(min_scores_create_time)) &Q(create_time__lte=(max_scores_create_time))).order_by('competence_id')
            else:
                qtEditScores=""
            assessment_id_list=qtEditScores.order_by('competence_id').values('assessment_id')
            for i in range(len(assessment_id_list)):
                standard_score = r_rules_position.objects.filter(Q(position_id=is_user.objects.filter(id=user_id)[0].position_id)& Q(rules_id=competence_id_list[i]))[0].standard_score
                # print(score_list[i])
                # print(standard_score)
                gap =int(score_list[i]) - int(standard_score)
                is_competition_score.objects.filter(assessment_id=int(assessment_id_list[i]['assessment_id'])).update(manager_assessment=score_list[i],manager_gap=gap,create_time=datetime.datetime.now())
            return HttpResponse(json.dumps({"info": "评分提交成功",
                                            "display": "0"}),
                                            content_type="application/json")
        else:
            #返回现有的所有测评规则以及数量
            positionID=is_user.objects.filter(id=user_id)[0].position_id
            qtIDRules=r_rules_position.objects.filter(position_id=positionID).order_by("rules_id")
            rule_ids=[]
            #自评分数
            self_list=[]
            for i in qtIDRules:
                if is_competition_rule.objects.filter(competence_id=i.rules_id).exists():
                    rule_ids.append(i.rules_id)
            competence_rules = is_competition_rule.objects.filter(competence_id__in=rule_ids).order_by('competence_id')
            num_rules = competence_rules.count()
            contentForm = judgeForm()
            #score到达此处肯定已经自评完成
            qtScores=is_competition_score.objects.filter(user_id=user_id)
            max_scores_create_time=(qtScores.order_by('-create_time')[:1])[0].create_time
            min_scores_create_time=max_scores_create_time+datetime.timedelta(seconds=-4)
            #切分为两部分，一部分为只可以查看的editScores，另一部分只可以查看readScores
            if qtScores.filter(Q(create_time__gte=(min_scores_create_time)) &Q(create_time__lte=(max_scores_create_time))).exists():
                qtEditScores=qtScores.filter(Q(create_time__gte=(min_scores_create_time)) &Q(create_time__lte=(max_scores_create_time))).order_by('competence_id')
            else:
                qtEditScores=""
            for i in qtEditScores:
                if is_competition_score.objects.filter(assessment_id=i.assessment_id).exists():
                    self_list.append(i.self_assessment)
            if is_group_leader == 1:
                is_manager = 1
            feedbackInfo={"is_staff":is_staff,"is_manager":is_manager,"is_departmentManager":is_departmentManager,"is_group_leader":is_group_leader,
                        "is_plantManager":is_plantManager,"num_rules":num_rules,'competence_id':rule_ids,"user_id":user_id,"self_assessment":self_list}
            return render(request,'QtCompetence/competence_judge.html',{'contentForm':contentForm,
                                                                        'competence_rules':competence_rules,
                                                                        "feedbackInfo": json.dumps(feedbackInfo),
                                                                        "real_name":real_name,
                                                                        "is_staff":is_staff,
                                                                        "user_pk":user_pk,
                                                                        "is_manager":is_manager,
                                                                        "is_departmentManager":is_departmentManager,
                                                                        "is_plantManager":is_plantManager,
                                                                        "is_group_leader":is_group_leader})
    else:
        return redirect('QtCompetence:login_user')

#经理管理员工评价状况
class competence_manage(ListView):
    model = is_user
    template_name = 'QtCompetence/competence_manage.html'
    context_object_name='user_list'

    def get_context_data(self, **kwargs):
        not_self_judge_list=[]
        manager_judge_list=[]
        context = super().get_context_data(**kwargs)
        real_name = self.request.session.get('real_name', None)
        user_pk = self.request.session.get('user_pk', None)
        is_staff = self.request.session.get('is_staff',None)
        is_manager = self.request.session.get('is_manager',None)
        is_departmentManager = self.request.session.get('is_departmentManager',None)
        is_plantManager = self.request.session.get('is_plantManager',None)
        is_group_leader = self.request.session.get('is_group_leader',None)
        context['real_name'] = real_name
        context['is_staff']=is_staff
        context['user_pk']=user_pk
        context['is_manager']=is_manager
        context['is_departmentManager']=is_department
        context['is_plantManager']=is_plantManager
        context['is_group_leader']=is_group_leader
        #未自评员工
        if r_level.objects.filter(L1_id=user_pk).exists():
            qtLevel_id=r_level.objects.filter(L1_id=user_pk)
            for i in qtLevel_id:
                if is_competition_score.objects.filter(user_id=i.user_id).exists():
                    #查看此处最新的29条记录是否被manager评价
                    qtScores=is_competition_score.objects.filter(user_id=i.user_id)
                    max_scores_create_time=(qtScores.order_by('-create_time')[:1])[0].create_time
                    min_scores_create_time=max_scores_create_time+datetime.timedelta(seconds=-4)
                    if qtScores.filter(Q(create_time__gte=(min_scores_create_time)) &Q(create_time__lte=(max_scores_create_time))).exists():
                        qtEditScores=qtScores.filter(Q(create_time__gte=(min_scores_create_time)) &Q(create_time__lte=(max_scores_create_time))).order_by('-create_time')
                    else:
                        qtEditScores=""
                    if qtEditScores[0].manager_assessment:
                        manager_judge_list.append(i.user_id)
                    else:
                        pass
                else:
                    not_self_judge_list.append(i.user_id)
        not_self_judge = is_user.objects.filter(id__in=not_self_judge_list)
        manager_judge = is_user.objects.filter(id__in=manager_judge_list)
        scores = (is_competition_score.objects.filter(user_id__in=manager_judge_list).extra(select={"create_time": "DATE_FORMAT(create_time, '%%Y-%%m-%%d %%H:%%i')"}).values('user_id','create_time')).distinct().order_by('user_id')
        #scores = (is_competition_score.objects.filter(user_id__in=manager_judge_list).values('user_id','create_time')).distinct().order_by('user_id')
        context['not_self_judge']=not_self_judge
        context['manager_judge']=manager_judge
        if not user_pk:
            return redirect('QtCompetence:login_user')
        #拼接不上去目前先暂时放着在这里
        #context['scores']=scores
        return context

    def get_queryset(self):
        #已自评员工
        self_judge_list=[]
        user_pk = self.request.session.get('user_pk', None)
        if r_level.objects.filter(L1_id=user_pk).exists():
            qtLevel_id=r_level.objects.filter(L1_id=user_pk)
            for i in qtLevel_id:
                if is_competition_score.objects.filter(user_id=i.user_id).exists():
                    #查看此处最新的29条记录是否被manager评价
                    qtScores=is_competition_score.objects.filter(user_id=i.user_id)
                    max_scores_create_time=(qtScores.order_by('-create_time')[:1])[0].create_time
                    min_scores_create_time=max_scores_create_time+datetime.timedelta(seconds=-4)
                    if qtScores.filter(Q(create_time__gte=(min_scores_create_time)) &Q(create_time__lte=(max_scores_create_time))).exists():
                        qtEditScores=qtScores.filter(Q(create_time__gte=(min_scores_create_time)) &Q(create_time__lte=(max_scores_create_time))).order_by('-create_time')
                    else:
                        qtEditScores=""
                    if not qtEditScores[0].manager_assessment:
                        self_judge_list.append(i.user_id)
                    else:
                        pass
        self_judge = is_user.objects.filter(id__in=self_judge_list)
        return self_judge

def competence_analyze(request):
    user_pos_list = []
    feedbackInfo = []
    feedbackInfoSelf1 = []
    feedbackInfoSelf2 = []
    feedbackInfoManager1 = []
    feedbackInfoManager2 = []
    user_pk = request.session.get('user_pk',None)
    user_name = request.session.get('user_name',None)
    real_name = request.session.get('real_name',None)
    is_staff = request.session.get('is_staff',None)
    is_manager = request.session.get('is_manager',None)
    is_departmentManager = request.session.get('is_departmentManager',None)
    is_plantManager = request.session.get('is_plantManager',None)
    is_group_leader = request.session.get('is_group_leader',None)
    if user_pk:
        if request.is_ajax():
            year=request.POST.get('year',None)
            month=request.POST.get('month',None)
            if year is not None and month is not None:
                pass
            else:
                time=is_competition_score.objects.extra(select={"create_time": "DATE_FORMAT(create_time, '%%Y-%%m')"}).values_list('create_time',flat=True).distinct()[0]
                year=time.split('-')[0]
                month=time.split('-')[1]
            #feedback1各项评价规则gap值合计
            sum_self_gap=0
            sum_manager_gap=0
            total_sum_self_gap=0
            total_sum_manager_gap=0
            #获取最新时间 初步规定一次测评周期为一个月
            qt_rule_names=is_competition_rule.objects.all()
            #每一条rule对应的gap<0的绝对值求总和
            for rule in qt_rule_names:
                qt_sum_one_rule=is_competition_score.objects.filter(Q(competence_id=rule.competence_id)&Q(create_time__year=year)&Q(create_time__month=month))
                if qt_sum_one_rule:
                    #只自评未经理评价
                    for score in qt_sum_one_rule:
                        if score.self_gap is not None and score.manager_gap is None:
                            if int(score.self_gap) < 0:
                                sum_self_gap += abs(int(score.self_gap))
                                sum_manager_gap += 0
                            else:
                                sum_self_gap += 0
                                sum_manager_gap += 0
                        else:    
                            if int(score.manager_gap) < 0:
                                sum_manager_gap += abs(int(score.manager_gap))
                            if int(score.self_gap) < 0:
                                sum_self_gap += abs(int(score.self_gap))
                else:
                    pass
                if sum_self_gap is not 0:
                    data={"name":rule.competence_name,"y":sum_self_gap}
                    feedbackInfoSelf1.append(data)
                else:
                    pass
                if sum_manager_gap is not 0:
                    data={"name":rule.competence_name,"y":sum_manager_gap}
                    feedbackInfoManager1.append(data)
                else:
                    pass
                sum_self_gap=0
                sum_manager_gap=0
            feedbackInfoSelf1.append({"name":"others","y":0})
            feedbackInfoManager1.append({"name":"others","y":0})
            #feedback2各职位gap值合计
            #获取当前登录人的部门
            qtUser=is_user.objects.filter(id=user_pk)
            qt_user_pos=is_user.objects.filter(department_id=qtUser[0].department_id)
            for user in qt_user_pos:
                user_pos_list.append(user.position_id)
            #质量部对应所有岗位
            qt_pos=is_position.objects.filter(id__in=user_pos_list)
            #pos_id获取当前下的所有员工
            for pos in qt_pos:
                rule_id_list=[]
                sum_one_pos=0
                qt_user=is_user.objects.filter(position_id=pos.id)
                qt_rule=r_rules_position.objects.filter(position_id=pos.id)
                for rule in qt_rule:
                    qt_scores=is_competition_score.objects.filter(competence_id=rule.rules_id)
                    #一条规则得分合计是该规则对应所有应测评人的gap值合计
                    #一个岗位多个人
                    for user in qt_user:
                        sum_self_gap=0
                        sum_manager_gap=0
                        qt_sum_dept_score=is_competition_score.objects.filter(Q(competence_id=rule.rules_id)&Q(user_id=user.id)&Q(create_time__year=year)&Q(create_time__month=month))
                        if qt_sum_dept_score:
                            for score in qt_sum_dept_score:
                                if score.self_gap is not None and score.manager_gap is None:
                                    if int(score.self_gap) < 0:
                                        sum_self_gap += abs(int(score.self_gap))
                                        sum_manager_gap += 0
                                    else:
                                        sum_self_gap += 0
                                        sum_manager_gap += 0
                                else:    
                                    if int(score.manager_gap) < 0:
                                        sum_manager_gap += abs(int(score.manager_gap))
                                    if int(score.self_gap) < 0:
                                        sum_self_gap += abs(int(score.self_gap))
                        else:
                            pass
                        total_sum_self_gap+=sum_self_gap
                        total_sum_manager_gap+=sum_manager_gap
                if total_sum_self_gap is not 0:
                    data={"name":pos.position_name,"y":total_sum_self_gap}
                    feedbackInfoSelf2.append(data)
                else:
                    pass
                if total_sum_manager_gap is not 0:
                    data={"name":pos.position_name,"y":total_sum_manager_gap}
                    feedbackInfoManager2.append(data)
                else:
                    pass
                sum_self_gap=0
                sum_manager_gap=0
                total_sum_self_gap=0
                total_sum_manager_gap=0
            feedbackInfoSelf2.append({"name":"others","y":0})
            feedbackInfoManager2.append({"name":"others","y":0})
            feedbackInfo.append(feedbackInfoSelf1)
            feedbackInfo.append(feedbackInfoSelf2)
            feedbackInfo.append(feedbackInfoManager1)
            feedbackInfo.append(feedbackInfoManager2)
            feedbackInfo.append({"year":year,"month":month})
            return HttpResponse(json.dumps({"feedbackInfo": feedbackInfo}),
                                            content_type="application/json")
        else:
            if is_competition_score.objects.values_list('create_time').exists():
                return render(request,'QtCompetence/competence_analyze.html',{"selectform":selectTime,
                                                                            "real_name":real_name,
                                                                            "is_staff":is_staff,
                                                                            "user_pk":user_pk,
                                                                            "is_manager":is_manager,
                                                                            "is_departmentManager":is_departmentManager,
                                                                            "is_plantManager":is_plantManager,
                                                                            "is_group_leader":is_group_leader})
            else:
                 return render(request,'QtCompetence/competence_analyze.html',{"selectform":"",
                                                                            "real_name":real_name,
                                                                            "is_staff":is_staff,
                                                                            "user_pk":user_pk,
                                                                            "is_manager":is_manager,
                                                                            "is_departmentManager":is_departmentManager,
                                                                            "is_plantManager":is_plantManager,
                                                                            "is_group_leader":is_group_leader})
    else:
        return redirect('QtCompetence:login_user')
    
def audit_analyze(request):
    user_pos_list = []
    feedbackInfo = []
    #Conforming Ratio
    feedbackInfo1 = []
    #Recurring Ratio
    feedbackInfo2 = []
    #time
    feedbackTime = []
    #Non-confirming qty
    feedbackInfo3 = []
    #Recurring qty
    feedbackInfo4 = []
    #Non-confirming target
    feedbackInfo5 = [86,86,86,86,86,86,86,86,86,86,86,86]
    #Recurring target
    feedbackInfo6 = [86,86,86,86,86,86,86,86,86,86,86,86]
    user_pk = request.session.get('user_pk',None)
    user_name = request.session.get('user_name',None)
    real_name = request.session.get('real_name',None)
    is_staff = request.session.get('is_staff',None)
    is_manager = request.session.get('is_manager',None)
    is_departmentManager = request.session.get('is_departmentManager',None)
    is_plantManager = request.session.get('is_plantManager',None)
    is_group_leader = request.session.get('is_group_leader',None)
    if user_pk:
        if request.is_ajax():
            year=request.POST.get('year',None)
            month=request.POST.get('month',None)
            if year is not None and month is not None:
                pass
            else:
                time=is_audit.objects.extra(select={"audit_date": "DATE_FORMAT(audit_date, '%%Y-%%m')"}).values_list('audit_date',flat=True).distinct()[0]
                year=time.split('-')[0]
                month=time.split('-')[1]     
            #针对所有IATF进行统计
            #qt_IATF=is_IATF.objects.all()
            # for rule in qt_IATF:
            #     #对应168IATF
            #     #IATF对应IATF_detail
            #     qt_IATF_detail=is_IATF_detail.objects.filter(IATF_id=rule.IATF_id)
            #     if qt_IATF_detail:
            #         IATF_detail_list=[] 
            #         for i in qt_IATF_detail:
            #             IATF_detail_list.append(i.IATF_detail_id)
            #         IATF_detail=qt_IATF_detail.values_list('IATF_detail_id')      
            #         time=datetime.datetime(int(year),int(month),28,23,59,59)
            #         qt_fingding=is_finding.objects.filter(create_time__range=[time-datetime.timedelta(days=365),time],IATF_detail_id__in=IATF_detail_list)
            #         print(qt_fingding)
            #         if qt_fingding:
            #             #去年某月到今年十二个月的整体
            #             #重复出现过则出发次数加一
            #             sum_if_IATF += 1
            #             #共出现过几次
            #             sum_if_IATF_times += len(qt_fingding)
            #  就差大于四的部分
            # 对应699 detail IATF直接筛选比较
            qt_IATF_detail=is_IATF_detail.objects.all()
            if qt_IATF_detail:
                for i in range(0,12):
                    #是否触发168
                    sum_if_IATF=0
                    #触发168频率
                    sum_if_IATF_times=0
                    #是否触发次数大于2
                    sum_times_gt_two=0
                    #是否触发次数大于2数量
                    sum_qty_gt_two=0
                    #不符合项数目
                    non_confirming_qty=0
                    #一年十二个月循环
                    #拼接十二个月的结果
                    conforming_ratio_list=[]
                    recurring_ratio_list=[]
                    #recurring qty
                    recrurring_qty_list=[]
                    #non-confirming qty
                    non_confirming_list=[]                   
                    for rule in qt_IATF_detail:
                        time=datetime.datetime(int(year),int(month),28,23,59,59)
                        a=time-datetime.timedelta(days=365)-datetime.timedelta(days=30*i)
                        qt_fingding=is_finding.objects.filter(IATF_detail_id=rule.IATF_detail_id,create_time__range=(time-datetime.timedelta(days=365)-datetime.timedelta(days=30*i),time-datetime.timedelta(days=30*i)))
                        # print(qt_fingding)
                        if qt_fingding:
                            #去年某月到今年十二个月的整体
                            #重复出现过则触发次数加一
                            sum_if_IATF += 1
                            print(sum_if_IATF)
                            #共出现过几次
                            sum_if_IATF_times += len(qt_fingding)
                            print(sum_if_IATF_times)
                            non_confirming_qty=len(qt_fingding)
                        if len(qt_fingding) > 2:
                            sum_times_gt_two += 1
                            sum_qty_gt_two += len(qt_fingding)
                            print(sum_times_gt_two)
                    conforming_ratio=((len(qt_IATF_detail)-sum_if_IATF)/len(qt_IATF_detail))*100
                    print(conforming_ratio)
                    recurring_ratio=(sum_times_gt_two/len(qt_IATF_detail))*100
                    print(recurring_ratio)
                    conforming_ratio_list.append(conforming_ratio)
                    recurring_ratio_list.append(recurring_ratio)
                    recrurring_qty_list.append(sum_qty_gt_two)
                    non_confirming_list.append(non_confirming_qty)
                    feedbackInfo1.append(conforming_ratio_list)
                    feedbackInfo2.append(recurring_ratio_list)
                    feedbackInfo3.append(non_confirming_qty)
                    feedbackInfo4.append(recrurring_qty_list)
            #时间
            for i in range(0,12):
                if int(month)-i==0:
                    feedbackTime.insert(0,'Dec')
                elif int(month)-i==-1:
                    feedbackTime.insert(0,'Nov')
                elif int(month)-i==-2:
                    feedbackTime.insert(0,'Oct')
                elif int(month)-i==-3:
                    feedbackTime.insert(0,'Sep')
                elif int(month)-i==-4:
                    feedbackTime.insert(0,'Aug')
                elif int(month)-i==-5:
                    feedbackTime.insert(0,'Jul')
                elif int(month)-i==-6:
                    feedbackTime.insert(0,'Jun')
                elif int(month)-i==-7:
                    feedbackTime.insert(0,'May')
                elif int(month)-i==-8:
                    feedbackTime.insert(0,'Apr')
                elif int(month)-i==-9:
                    feedbackTime.insert(0,'Mar')
                elif int(month)-i==-10:
                    feedbackTime.insert(0,'Feb')
                elif int(month)-i==-11:
                    feedbackTime.insert(0,'Jan')
                elif int(month)-i==1:
                    feedbackTime.insert(0,'Jan')
                elif int(month)-i==2:
                    feedbackTime.insert(0,'Feb')
                elif int(month)-i==3:
                    feedbackTime.insert(0,'Mar')
                elif int(month)-i==4:
                    feedbackTime.insert(0,'Apr')
                elif int(month)-i==5:
                    feedbackTime.insert(0,'May')
                elif int(month)-i==6:
                    feedbackTime.insert(0,'Jun')
                elif int(month)-i==7:
                    feedbackTime.insert(0,'Jul')
                elif int(month)-i==8:
                    feedbackTime.insert(0,'Aug')
                elif int(month)-i==9:
                    feedbackTime.insert(0,'Sep')
                elif int(month)-i==10:
                    feedbackTime.insert(0,'Oct')
                elif int(month)-i==11:
                    feedbackTime.insert(0,'Nov')
                elif int(month)-i==12:
                    feedbackTime.insert(0,'Dec')
            feedbackInfo.append(feedbackInfo1)
            feedbackInfo.append(feedbackInfo2)
            feedbackInfo.append(feedbackInfo3)
            feedbackInfo.append(feedbackInfo4)
            feedbackInfo.append(feedbackInfo5)
            feedbackInfo.append(feedbackInfo6)
            feedbackInfo.append(feedbackTime)
            return HttpResponse(json.dumps({"feedbackInfo": feedbackInfo}),
                                            content_type="application/json")
        else:
            if is_finding.objects.values_list("create_time").exists():
                return render(request,'QtCompetence/audit_analyze.html',{"selectFindingTime":selectFindingTime,
                                                                        "real_name":real_name,
                                                                        "is_staff":is_staff,
                                                                        "user_pk":user_pk,
                                                                        "is_manager":is_manager,
                                                                        "is_departmentManager":is_departmentManager,
                                                                        "is_plantManager":is_plantManager,
                                                                        "is_group_leader":is_group_leader})
            else:
                return render(request,'QtCompetence/audit_analyze.html',{"selectFindingTime":"",
                                                                        "real_name":real_name,
                                                                        "is_staff":is_staff,
                                                                        "user_pk":user_pk,
                                                                        "is_manager":is_manager,
                                                                        "is_departmentManager":is_departmentManager,
                                                                        "is_plantManager":is_plantManager,
                                                                        "is_group_leader":is_group_leader})
    else:
        return redirect('QtCompetence:login_user')

def personal_home(request):
    user_pos_list = []
    feedbackInfo = []
    feedbackInfoSelf1 = []
    feedbackInfoSelf2 = []
    feedbackInfoManager1 = []
    feedbackInfoManager2 = []
    user_pk = request.session.get('user_pk',None)
    user_name = request.session.get('user_name',None)
    real_name = request.session.get('real_name',None)
    is_staff = request.session.get('is_staff',None)
    is_manager = request.session.get('is_manager',None)
    is_departmentManager = request.session.get('is_departmentManager',None)
    is_plantManager = request.session.get('is_plantManager',None)
    is_group_leader = request.session.get('is_group_leader',None)
    if user_pk:
        if is_staff:
            return redirect(reverse('QtCompetence:staff_home', args=(user_pk,)))
        if is_manager or is_group_leader:
            if request.is_ajax():
                year=request.POST.get('year',None)
                month=request.POST.get('month',None)
                if year is not None and month is not None:
                    pass
                else:
                    time=is_competition_score.objects.extra(select={"create_time": "DATE_FORMAT(create_time, '%%Y-%%m')"}).values_list('create_time',flat=True).distinct()[0]
                    year=time.split('-')[0]
                    month=time.split('-')[1]
                #feedback1各项评价规则gap值合计
                sum_self_gap=0
                sum_manager_gap=0
                total_sum_self_gap=0
                total_sum_manager_gap=0
                #获取最新时间 初步规定一次测评周期为一个月
                #筛选出manager下所有员工对应的考评项目
                #1.筛选出manger下所有员工的id对应的position id
                temp_pos_list=[]
                temp_rule_id_list=[]
                if r_level.objects.filter(L1_id=user_pk).exists():
                    qtLevel_id=r_level.objects.filter(L1_id=user_pk)
                    for i in qtLevel_id:
                        temp_pos_list.append(is_user.objects.filter(id=i.user_id)[0].position_id)
                #去重
                #print(temp_pos_list)
                pos_list=set(temp_pos_list)
                #print(pos_list)
                #每个岗位对应的competence_id
                qtRules=r_rules_position.objects.filter(position_id__in=pos_list)
                for i in qtRules:
                    temp_rule_id_list.append(i.rules_id)
                #print(temp_rule_id_list)
                rule_id_list=set(temp_rule_id_list)
                #print(rule_id_list)
                qt_rule_names=is_competition_rule.objects.filter(competence_id__in=rule_id_list)
                #清空
                temp_pos_list=[]
                temp_rule_id_list=[]
                #每一条rule对应的gap<0的绝对值求总和
                for rule in qt_rule_names:
                    qt_sum_one_rule=is_competition_score.objects.filter(Q(competence_id=rule.competence_id)&Q(create_time__year=year)&Q(create_time__month=month))
                    if qt_sum_one_rule:
                        #只自评未经理评价
                        for score in qt_sum_one_rule:
                            if score.self_gap is not None and score.manager_gap is None:
                                if int(score.self_gap) < 0:
                                    sum_self_gap += abs(int(score.self_gap))
                                    sum_manager_gap += 0
                                else:
                                    sum_self_gap += 0
                                    sum_manager_gap += 0
                            else:    
                                if int(score.manager_gap) < 0:
                                    sum_manager_gap += abs(int(score.manager_gap))
                                if int(score.self_gap) < 0:
                                    sum_self_gap += abs(int(score.self_gap))
                    else:
                        pass
                    if sum_self_gap is not 0:
                        data={"name":rule.competence_name,"y":sum_self_gap}
                        feedbackInfoSelf1.append(data)
                    else:
                        pass
                    if sum_manager_gap is not 0:
                        data={"name":rule.competence_name,"y":sum_manager_gap}
                        feedbackInfoManager1.append(data)
                    else:
                        pass
                    sum_self_gap=0
                    sum_manager_gap=0
                feedbackInfoSelf1.append({"name":"others","y":0})
                feedbackInfoManager1.append({"name":"others","y":0})
                #feedback2各职位gap值合计
                #1.筛选出manger下所有员工的id对应的position id
                temp_pos_list=[]
                if r_level.objects.filter(L1_id=user_pk).exists():
                    qtLevel_id=r_level.objects.filter(L1_id=user_pk)
                    for i in qtLevel_id:
                        temp_pos_list.append(is_user.objects.filter(id=i.user_id)[0].position_id)
                #去重
                #print(temp_pos_list)
                pos_list=set(temp_pos_list)
                temp_pos_list=[]
                #print(pos_list)
                qt_pos=is_position.objects.filter(id__in=pos_list)
                #pos_id获取当前下的所有员工
                for pos in qt_pos:
                    rule_id_list=[]
                    sum_one_pos=0
                    qt_user=is_user.objects.filter(position_id=pos.id)
                    qt_rule=r_rules_position.objects.filter(position_id=pos.id)
                    for rule in qt_rule:
                        qt_scores=is_competition_score.objects.filter(competence_id=rule.rules_id)
                        #一条规则得分合计是该规则对应所有应测评人的gap值合计
                        #一个岗位多个人
                        for user in qt_user:
                            sum_self_gap=0
                            sum_manager_gap=0
                            qt_sum_dept_score=is_competition_score.objects.filter(Q(competence_id=rule.rules_id)&Q(user_id=user.id)&Q(create_time__year=year)&Q(create_time__month=month))
                            if qt_sum_dept_score:
                                for score in qt_sum_dept_score:
                                    if score.self_gap is not None and score.manager_gap is None:
                                        if int(score.self_gap) < 0:
                                            sum_self_gap += abs(int(score.self_gap))
                                            sum_manager_gap += 0
                                        else:
                                            sum_self_gap += 0
                                            sum_manager_gap += 0
                                    else:    
                                        if int(score.manager_gap) < 0:
                                            sum_manager_gap += abs(int(score.manager_gap))
                                        if int(score.self_gap) < 0:
                                            sum_self_gap += abs(int(score.self_gap))
                            else:
                                pass
                            total_sum_self_gap+=sum_self_gap
                            total_sum_manager_gap+=sum_manager_gap
                    if total_sum_self_gap is not 0:
                        data={"name":pos.position_name,"y":total_sum_self_gap}
                        feedbackInfoSelf2.append(data)
                    else:
                        pass
                    if total_sum_manager_gap is not 0:
                        data={"name":pos.position_name,"y":total_sum_manager_gap}
                        feedbackInfoManager2.append(data)
                    else:
                        pass
                    sum_self_gap=0
                    sum_manager_gap=0
                    total_sum_self_gap=0
                    total_sum_manager_gap=0
                feedbackInfoSelf2.append({"name":"others","y":0})
                feedbackInfoManager2.append({"name":"others","y":0})
                feedbackInfo.append(feedbackInfoSelf1)
                feedbackInfo.append(feedbackInfoSelf2)
                feedbackInfo.append(feedbackInfoManager1)
                feedbackInfo.append(feedbackInfoManager2)
                feedbackInfo.append({"year":year,"month":month})
                return HttpResponse(json.dumps({"feedbackInfo": feedbackInfo}),
                                                content_type="application/json")
            else:
                user_list=[]
                manager_feedback = []
                #未自评员工
                if r_level.objects.filter(L1_id=user_pk).exists():
                    qtLevel_id=r_level.objects.filter(L1_id=user_pk)
                    for i in qtLevel_id:
                        user_list.append(i.user_id)
                qtScores = is_competition_score.objects.filter(user_id__in=user_list)
                return render(request,'QtCompetence/personal_home.html',{"selectform":selectTime,
                                                                            "real_name":real_name,
                                                                            "is_staff":is_staff,
                                                                            "is_manager":is_manager,
                                                                            "is_departmentManager":is_departmentManager,
                                                                            "is_plantManager":is_plantManager,
                                                                            "is_group_leader":is_group_leader,
                                                                            "user_pk":user_pk,
                                                                            "manager_feedback":qtScores})
        if is_plantManager or is_departmentManager:
            return redirect('QtCompetence:competence_analyze')
            
    else:
        return redirect('QtCompetence:login_user')
def staff_home(request,user_id):
    user_pk = request.session.get('user_pk',None)
    user_name = request.session.get('user_name',None)
    real_name = request.session.get('real_name',None)
    is_staff = request.session.get('is_staff',None)
    is_manager = request.session.get('is_manager',None)
    is_departmentManager = request.session.get('is_departmentManager',None)
    is_plantManager = request.session.get('is_plantManager',None)
    is_group_leader = request.session.get('is_group_leader',None)
    feedbackInfo = []
    feedbackInfoSelf1 = []
    feedbackInfoManager1 = []
    if user_pk:
        if request.is_ajax():
            year=request.POST.get('year',None)
            month=request.POST.get('month',None)
            if year is not None and month is not None:
                pass
            else:
                time=is_competition_score.objects.extra(select={"create_time": "DATE_FORMAT(create_time, '%%Y-%%m')"}).values_list('create_time',flat=True).distinct()[0]
                year=time.split('-')[0]
                month=time.split('-')[1]
            #feedback1各项评价规则gap值合计
            #获取最新时间 初步规定一次测评周期为一个月
            #每个岗位对应的competence_id
            rule_id_list=[]
            qtRules=r_rules_position.objects.filter(position_id=is_user.objects.filter(id=user_id)[0].position_id)
            for i in qtRules:
                rule_id_list.append(i.rules_id)
            qt_rule_names=is_competition_rule.objects.filter(competence_id__in=rule_id_list)
            for rule in qt_rule_names:
                qt_sum_one_rule=is_competition_score.objects.filter(Q(competence_id=rule.competence_id)&Q(create_time__year=year)&Q(create_time__month=month)&Q(user_id=user_id))
                if qt_sum_one_rule:
                    if qt_sum_one_rule[0].self_gap<0:
                        data={"name":rule.competence_name,"y":abs(int(qt_sum_one_rule[0].self_gap))}
                        feedbackInfoSelf1.append(data)
                    if qt_sum_one_rule[0].manager_gap is None:
                        data={"name":rule.competence_name,"y":0}
                        feedbackInfoManager1.append(data)
                    elif qt_sum_one_rule[0].manager_gap < 0:
                        data={"name":rule.competence_name,"y":abs(int(qt_sum_one_rule[0].manager_gap))}
                        feedbackInfoManager1.append(data)
            feedbackInfoSelf1.append({"name":"others","y":0})
            feedbackInfoManager1.append({"name":"others","y":0})
            feedbackInfo.append(feedbackInfoSelf1)
            feedbackInfo.append(feedbackInfoManager1)
            feedbackInfo.append({"year":year,"month":month})
            return HttpResponse(json.dumps({"feedbackInfo": feedbackInfo}),
                                            content_type="application/json")
        else:
            qtScores=is_competition_score.objects.filter(user_id=user_pk)
            #获取某个人所有评价日期的不重复值
            #all_judge_create_times = qtScores.values('create_time').distinct()
            #qtAllJudgeCreateTimes=(qtScores.extra(select={"create_time": "DATE_FORMAT(create_time, '%%Y-%%m-%%d %%H:%%i')"}).values('create_time')).distinct().order_by('-create_time')
            #有一条记录的人
            max_scores_create_time=(qtScores.order_by('-create_time')[:1])[0].create_time
            min_scores_create_time=max_scores_create_time+datetime.timedelta(seconds=-4)
            #切分为两部分，一部分为只可以查看的editScores，另一部分只可以查看readScores
            if qtScores.filter(~(Q(create_time__gte=(min_scores_create_time)) &Q(create_time__lte=(max_scores_create_time)))).exists():
                qtReadScores=qtScores.filter(~(Q(create_time__gte=(min_scores_create_time)) &Q(create_time__lte=(max_scores_create_time)))).order_by('-create_time')
            else:
                qtReadScores=""
            if qtScores.filter(Q(create_time__gte=(min_scores_create_time)) &Q(create_time__lte=(max_scores_create_time))).exists():
                qtEditScores=qtScores.filter(Q(create_time__gte=(min_scores_create_time)) &Q(create_time__lte=(max_scores_create_time)))
            # if is_competition_score.objects.filter(user_id=user_id).exists():
            #     qtCompetence_score=is_competition_score.objects.filter(user_id=user_id)
            #     competence_id_list = ','.join([str(q.competence_id) for q in qtCompetence_score])
            #     competence_id_list = competence_id_list.split(',')
            #     competence_rules = is_competition_rule.objects.filter(competence_id__in=competence_id_list)
            #     staff_feedback=[]
            #     # 手动拼接
            #     for i in range(len(competence_rules)):
            #         feedback=[]
            #         feedback.append(competence_rules[i].competence_name)
            #         feedback.append(qtCompetence_score[i].self_assessment)
            #         feedback.append(qtCompetence_score[i].self_gap)
            #         feedback.append(qtCompetence_score[i].manager_assessment)
            #         feedback.append(qtCompetence_score[i].manager_gap)
            #         # 如果自评与经理评价均完成,个人改正措施为空，经理改正措施为空则需要进行改正措施添
            #         is_need_add_self_correction=qtCompetence_score[i].self_assessment is not None and qtCompetence_score[i].manager_assessment is not None and qtCompetence_score[i].self_correction is None and qtCompetence_score[i].manager_correction is None
            #         is_need_add_manager_correction=qtCompetence_score[i].self_assessment is not None and qtCompetence_score[i].manager_assessment is not None and qtCompetence_score[i].self_correction is not None and qtCompetence_score[i].manager_correction is None
            #         is_need_update_self_correction=qtCompetence_score[i].self_assessment is not None and qtCompetence_score[i].manager_assessment is not None and qtCompetence_score[i].self_correction is not None 
            #         is_need_update_manager_correction=qtCompetence_score[i].self_assessment is not None and qtCompetence_score[i].manager_assessment is not None and qtCompetence_score[i].self_correction is not None and qtCompetence_score[i].manager_correction is not None
            #         feedback.append(is_need_add_self_correction)
            #         feedback.append(is_need_add_manager_correction)
            #         feedback.append(is_need_update_self_correction)
            #         feedback.append(is_need_update_manager_correction)
            #         if is_need_add_self_correction:
            #             pass
            #         staff_feedback.append(feedback)
                return render(request,'QtCompetence/staff_home.html',{"selectform":selectTime,
                                                                    "real_name":real_name,
                                                                    "user_pk":user_pk,
                                                                    "is_staff":is_staff,
                                                                    "is_manager":is_manager,
                                                                    "is_departmentManager":is_departmentManager,
                                                                    "is_plantManager":is_plantManager,
                                                                    "qtEditScores":qtEditScores,
                                                                    "qtReadScores":qtReadScores,
                                                                    "is_group_leader":is_group_leader})
            else:
                return redirect('QtCompetence:competence_judge')
    else:
        return redirect('QtCompetence:login_user')        
def logout_user(request):
    user_pk = request.session.get('user_pk', None)
    user_name = request.session.get('user_name', None)
    user_pwd = request.session.get('user_pwd', None)
    real_name = request.session.get('real_name', None)
    is_staff = request.session.get('is_staff',None)
    is_manager = request.session.get('is_manager',None)
    is_departmentManager = request.session.get('is_departmentManager',None)
    is_plantManager = request.session.get('is_plantManager',None)
    is_group_leader = request.session.get('is_group_leader',None)
    if user_pk is not None:
        del request.session['user_pk']
    if user_name is not None:
        del request.session['user_name']
    if user_pwd is not None:
        del request.session['user_pwd']
    if real_name is not None:
        del request.session['real_name']
    if is_staff is not None:
        del request.session['is_staff']
    if is_manager is not None:
        del request.session['is_manager']
    if is_departmentManager is not None:
        del request.session['is_departmentManager']
    if is_plantManager is not None:
        del request.session['is_plantManager']
    if is_group_leader is not None:
        del request.session['is_group_leader']
    return redirect('QtCompetence:home')


#Audit
def audit_home(request):
    user_pk = request.session.get('user_pk',None)
    if not user_pk:
        return redirect('QtCompetence:login_user')
    else:
        user_name = request.session.get('user_name',None)
        real_name = request.session.get('real_name',None)
        is_staff = request.session.get('is_staff',None)
        is_manager = request.session.get('is_manager',None)
        is_departmentManager = request.session.get('is_departmentManager',None)
        is_plantManager = request.session.get('is_plantManager',None)
        is_group_leader = request.session.get('is_group_leader',None)
        return render(request,'QtCompetence/audit_home.html',{"real_name":real_name,
                                                            "user_pk":user_pk,
                                                            "is_staff":is_staff,
                                                            "is_manager":is_manager,
                                                            "is_departmentManager":is_departmentManager,
                                                            "is_plantManager":is_plantManager,
                                                            "is_group_leader":is_group_leader})                                                       

#级联选择audit_body_detail
def ajax_load_audit_body_detail(request):
    audit_body_id=request.GET.get('audit_body_id',None)
    audit_body_detail = is_audit_body_detail.objects.filter(audit_body_id=audit_body_id).order_by('audit_body_detail_id')
    return render(request,'QtCompetence/audit_body_detail_dropdown_list_options.html',{'audit_body_detail':audit_body_detail})

#is_audit input
class audit_info_create(CreateView,PermissionRequiredMixin):
    template_name = 'QtCompetence/create_audit_with_dropdownlist.html'
    #model = is_audit
    success_url = reverse_lazy('QtCompetence:audit_info_list')
    #fields = ['audit_date', 'department', 'audit_body', 'audit_scope', 'audit_type','audit_purpose','user']
    form_class=auditCreateForm
    permission_required = 'QtCompetence.add_is_audit'
    raise_exception = True

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        real_name = self.request.session.get('real_name', None)
        user_pk = self.request.session.get('user_pk', None)
        is_staff = self.request.session.get('is_staff',None)
        is_manager = self.request.session.get('is_manager',None)
        is_departmentManager = self.request.session.get('is_departmentManager',None)
        is_plantManager = self.request.session.get('is_plantManager',None)
        is_group_leader = self.request.session.get('is_group_leader',None)
        context['real_name'] = real_name
        context['is_staff']=is_staff
        context['user_pk']=user_pk
        context['is_manager']=is_manager
        context['is_departmentManager']=is_department
        context['is_plantManager']=is_plantManager
        context['is_group_leader']=is_group_leader
        return context

    def post(self, request, *args, **kwargs):
        user_pk = self.request.session.get('user_pk', None)
        if user_pk:
            department = request.POST.get("department", None)
            audit_date = request.POST.get("audit_date", None)
            audit_body = request.POST.get("audit_body", None)
            audit_body_detail = request.POST.get("audit_body_detail", None)
            audit_scope = request.POST.get("audit_scope", None)
            audit_type = request.POST.get("audit_type", None)
            line = request.POST.get("line", None)
            user = request.POST.get("user", None)
            is_audit.objects.create(department_id=department,audit_body_id=audit_body,audit_body_detail_id=audit_body_detail,audit_scope=audit_scope,audit_type_id=audit_type,audit_date=audit_date,
                                    line=line,user_id=user)
            return redirect('QtCompetence:audit_info_list')
        else:
            return redirect('QtCompetence:login_user')

#is_audit update
class audit_info_update(UpdateView):
    template_name = 'QtCompetence/update_audit_with_dropdownlist.html'
    #model = is_audit
    success_url = reverse_lazy('QtCompetence:audit_info_list')
    #fields = ['audit_date', 'department', 'audit_body', 'audit_scope', 'audit_type','audit_purpose','host','clause_coverage']
    form_class=auditCreateForm
    queryset=is_audit.objects.all().order_by('audit_id')

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        real_name = self.request.session.get('real_name', None)
        user_pk = self.request.session.get('user_pk', None)
        is_staff = self.request.session.get('is_staff',None)
        is_manager = self.request.session.get('is_manager',None)
        is_departmentManager = self.request.session.get('is_departmentManager',None)
        is_plantManager = self.request.session.get('is_plantManager',None)
        is_group_leader = self.request.session.get('is_group_leader',None)
        context['real_name'] = real_name
        context['is_staff']=is_staff
        context['user_pk']=user_pk
        context['is_manager']=is_manager
        context['is_departmentManager']=is_department
        context['is_plantManager']=is_plantManager
        context['is_group_leader']=is_group_leader
        context['return_type']='audit_info'
        return context

    def post(self, request, *args, **kwargs):
        user_pk = self.request.session.get('user_pk', None)
        if user_pk:
            department = request.POST.get("department", None)
            audit_body = request.POST.get("audit_body", None)
            audit_body_detail = request.POST.get("audit_body_detail", None)
            audit_scope = request.POST.get("audit_scope", None)
            audit_type = request.POST.get("audit_type", None)
            line = request.POST.get("line", None)
            audit_date = request.POST.get("audit_date", None)
            user = request.POST.get("user", None)
            is_audit.objects.filter(audit_id=self.kwargs['pk']).update(department_id=department,audit_body_id=audit_body,audit_body_detail_id=audit_body_detail,audit_scope=audit_scope,audit_type_id=audit_type,audit_date=audit_date,
                                    line=line,user_id=user)
            return redirect('QtCompetence:audit_info_list')
        else:
            return redirect('QtCompetence:login_user')

#is_audit delete
class audit_info_delete(DeleteView):
    template_name = 'QtCompetence/delete.html'
    model = is_audit
    success_url = reverse_lazy('QtCompetence:audit_info_list')

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        real_name = self.request.session.get('real_name', None)
        user_pk = self.request.session.get('user_pk', None)
        is_staff = self.request.session.get('is_staff',None)
        is_manager = self.request.session.get('is_manager',None)
        is_departmentManager = self.request.session.get('is_departmentManager',None)
        is_plantManager = self.request.session.get('is_plantManager',None)
        is_group_leader = self.request.session.get('is_group_leader',None)
        context['real_name'] = real_name
        context['is_staff']=is_staff
        context['user_pk']=user_pk
        context['is_manager']=is_manager
        context['is_departmentManager']=is_department
        context['is_plantManager']=is_plantManager
        context['is_group_leader']=is_group_leader
        context['return_type']='audit_info'
        return context

    def post(self, request, *args, **kwargs):
        user_pk = self.request.session.get('user_pk', None)
        if user_pk:
            department = request.POST.get("department", None)
            audit_body = request.POST.get("audit_body", None)
            audit_scope = request.POST.get("audit_scope", None)
            audit_type = request.POST.get("audit_type", None)
            audit_purpose = request.POST.get("audit_purpose", None)
            user = request.POST.get("user", None)
            is_audit.objects.filter(audit_id=self.kwargs['pk']).delete()
            return redirect('QtCompetence:audit_info_list')
        else:
            return redirect('QtCompetence:login_user')

#is_audit list
class audit_info_list(ListView):
    #model = is_audit
    template_name = 'QtCompetence/audit_info_list.html'
    paginate_by = 20#一个页面显示的条目
    context_object_name = 'audit_info'

    def get_queryset(self):
        return audit_info_filter(self.request.GET,queryset=is_audit.objects.all().order_by("-audit_date")).qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        real_name = self.request.session.get('real_name', None)
        user_pk = self.request.session.get('user_pk', None)
        is_staff = self.request.session.get('is_staff',None)
        is_manager = self.request.session.get('is_manager',None)
        is_departmentManager = self.request.session.get('is_departmentManager',None)
        is_plantManager = self.request.session.get('is_plantManager',None)
        audit_filter = audit_info_filter(self.request.GET,queryset=is_audit.objects.all().order_by("-audit_date"))
        is_group_leader = self.request.session.get('is_group_leader',None)
        context['real_name'] = real_name
        context['is_staff']=is_staff
        context['user_pk']=user_pk
        context['is_manager']=is_manager
        context['is_departmentManager']=is_department
        context['is_plantManager']=is_plantManager
        context['is_group_leader']=is_group_leader
        context['audit_filter'] = audit_filter
        paginator = context.get('paginator')
        page_obj = context.get('page_obj')
        dept = self.request.GET.get('department','')
        user = self.request.GET.get('user','')
        audit_body = self.request.GET.get('audit_body','')
        paginator_data = self.get_pagination_data(paginator, page_obj,dept,user,audit_body)
        context.update(paginator_data) # 将当前字典的kv更新到context字典中。
        return context

    # 这里来负责跳转的页码处理
    def get_pagination_data(self, paginator, page_obj, dept='',user='',audit_body='',around_count=2): # arount_count=2表示从当前页前推两页，后推两页
        current_page = page_obj.number
        num_page = paginator.num_pages
        left_has_more = False # 左边还有没有未显示的页码
        right_has_more = False
        #判断当前页是不是比4小，比如当前页是第二页，他就不能存在 0.1.2.3.4这种情况。
        if current_page <= around_count + 2:
            left_page = range(1, current_page)
        else:
            left_has_more = True
            left_page = range(current_page-around_count, current_page)
        if current_page >= num_page-around_count-1:
            right_page = range(current_page+1, num_page+1)
        else:
            right_has_more = True
            right_page = range(current_page+1, current_page+3)
        return {
            'left_pages': left_page,
            'right_pages': right_page,
            'current_page': current_page,
            'left_has_more': left_has_more,
            'right_has_more': right_has_more,
            'dept':dept,
            'user':user,
            'audit_body':audit_body,
        }

#级联选择IATF_IATF_Deatil
def ajax_load_IATF_detail(request):
    IATF_Id=request.GET.get('IATF_Id',None)
    IATF_detail = is_IATF_detail.objects.filter(IATF_id=IATF_Id).order_by('IATF_detail_id')
    return render(request,'QtCompetence/IATF_detail_dropdown_list_options.html',{'IATF_detail':IATF_detail})

#is_finding input
class finding_info_create(CreateView):
    template_name = 'QtCompetence/create_finding_with_dropdownlist.html'
    #model = is_finding
    success_url = reverse_lazy('QtCompetence:finding_info_list')
    #fields = ['audit', 'finding_content', 'department', 'IATF_detail', 'weight','rootCause','correction']
    form_class=findingCreateForm

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        real_name = self.request.session.get('real_name', None)
        user_pk = self.request.session.get('user_pk', None)
        is_staff = self.request.session.get('is_staff',None)
        is_manager = self.request.session.get('is_manager',None)
        is_departmentManager = self.request.session.get('is_departmentManager',None)
        is_plantManager = self.request.session.get('is_plantManager',None)
        is_group_leader = self.request.session.get('is_group_leader',None)
        context['real_name'] = real_name
        context['is_staff']=is_staff
        context['user_pk']=user_pk
        context['is_manager']=is_manager
        context['is_departmentManager']=is_department
        context['is_plantManager']=is_plantManager
        context['is_group_leader']=is_group_leader
        return context

    def post(self, request, *args, **kwargs):
        user_pk = self.request.session.get('user_pk', None)
        if user_pk:
            audit = request.POST.get("audit", None)
            finding_content = request.POST.get("finding_content", None)
            audit_department = request.POST.get("audit_department", None)
            IATF = request.POST.get("IATF", None)
            IATF_detail = request.POST.get("IATF_detail", None)
            weight = request.POST.get("weight", None)
            user = request.POST.get("user", None)
            user_pk = self.request.session.get('user_pk', None)
            is_finding.objects.create(audit_department_id=audit_department,audit_id=audit,finding_content=finding_content,IATF_id=IATF,IATF_detail_id=IATF_detail,
                                    weight=weight,user_id=user_pk)
            return redirect('QtCompetence:finding_info_list')
        else:
            return redirect('QtCompetence:login_user')

#is_finding update
class finding_info_update(UpdateView):
    template_name = 'QtCompetence/update_finding_with_dropdownlist.html'
    #model = is_finding
    success_url = reverse_lazy('QtCompetence:finding_info_list')
    #fields = ['audit', 'finding_content', 'department', 'IATF_detail', 'weight','rootCause','correction']
    form_class=findingCreateForm
    queryset=is_finding.objects.all().order_by('finding_id')

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        real_name = self.request.session.get('real_name', None)
        user_pk = self.request.session.get('user_pk', None)
        is_staff = self.request.session.get('is_staff',None)
        is_manager = self.request.session.get('is_manager',None)
        is_departmentManager = self.request.session.get('is_departmentManager',None)
        is_plantManager = self.request.session.get('is_plantManager',None)
        is_group_leader = self.request.session.get('is_group_leader',None)
        context['real_name'] = real_name
        context['is_staff']=is_staff
        context['user_pk']=user_pk
        context['is_manager']=is_manager
        context['is_departmentManager']=is_department
        context['is_plantManager']=is_plantManager
        context['is_group_leader']=is_group_leader
        context['return_type']='finding_info'
        return context

    def post(self, request, *args, **kwargs):
        user_pk = self.request.session.get('user_pk', None)
        if user_pk:
            audit = request.POST.get("audit", None)
            finding_content = request.POST.get("finding_content", None)
            audit_department = request.POST.get("audit_department", None)
            IATF = request.POST.get("IATF", None)
            IATF_detail = request.POST.get("IATF_detail", None)
            weight = request.POST.get("weight", None)
            user = request.POST.get("user", None)
            user_pk = self.request.session.get('user_pk', None)
            is_finding.objects.filter(finding_id=self.kwargs['pk']).update(audit_department_id=audit_department,audit_id=audit,finding_content=finding_content,IATF_id=IATF,IATF_detail_id=IATF_detail,
                                    weight=weight,user_id=user_pk)
            return redirect('QtCompetence:finding_info_list')
        else:
            return redirect('QtCompetence:login_user')

#is_finding delete
class finding_info_delete(DeleteView):
    template_name = 'QtCompetence/delete.html'
    model = is_finding
    success_url = reverse_lazy('QtCompetence:finding_info_list')

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        real_name = self.request.session.get('real_name', None)
        user_pk = self.request.session.get('user_pk', None)
        is_staff = self.request.session.get('is_staff',None)
        is_manager = self.request.session.get('is_manager',None)
        is_departmentManager = self.request.session.get('is_departmentManager',None)
        is_plantManager = self.request.session.get('is_plantManager',None)
        is_group_leader = self.request.session.get('is_group_leader',None)
        context['real_name'] = real_name
        context['is_staff']=is_staff
        context['user_pk']=user_pk
        context['is_manager']=is_manager
        context['is_departmentManager']=is_department
        context['is_plantManager']=is_plantManager
        context['is_group_leader']=is_group_leader
        context['return_type']='finding_info'
        return context

    def post(self, request, *args, **kwargs):
        user_pk = self.request.session.get('user_pk', None)
        if user_pk:
            user_pk = self.request.session.get('user_pk', None)
            is_finding.objects.filter(finding_id=self.kwargs['pk']).delete()
            return redirect('QtCompetence:finding_info_list')
        else:
            return redirect('QtCompetence:login_user')

#is_finding list
class finding_info_list(ListView):
    #model = is_finding
    template_name = 'QtCompetence/finding_info_list.html'
    paginate_by = 20#一个页面显示的条目
    context_object_name = 'finding_info'

    def get_queryset(self):
        return finding_info_filter(self.request.GET,queryset=is_finding.objects.all().order_by("-audit__audit_date")).qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        real_name = self.request.session.get('real_name', None)
        user_pk = self.request.session.get('user_pk', None)
        is_staff = self.request.session.get('is_staff',None)
        is_manager = self.request.session.get('is_manager',None)
        is_departmentManager = self.request.session.get('is_departmentManager',None)
        is_plantManager = self.request.session.get('is_plantManager',None)
        finding_filter = finding_info_filter(self.request.GET,queryset=is_finding.objects.all().order_by("-audit__audit_date"))
        is_group_leader = self.request.session.get('is_group_leader',None)
        context['real_name'] = real_name
        context['is_staff']=is_staff
        context['user_pk']=user_pk
        context['is_manager']=is_manager
        context['is_departmentManager']=is_department
        context['is_plantManager']=is_plantManager
        context['is_group_leader']=is_group_leader
        context['finding_filter'] = finding_filter
        paginator = context.get('paginator')
        page_obj = context.get('page_obj')
        finding_content = self.request.GET.get('finding_content','')
        audit_department = self.request.GET.get('audit_department','')
        paginator_data = self.get_pagination_data(paginator, page_obj ,finding_content ,audit_department)
        context.update(paginator_data) # 将当前字典的kv更新到context字典中。
        return context

    # 这里来负责跳转的页码处理
    def get_pagination_data(self, paginator, page_obj, finding_content='', audit_department='', around_count=2): # arount_count=2表示从当前页前推两页，后推两页
        current_page = page_obj.number
        num_page = paginator.num_pages
        left_has_more = False # 左边还有没有未显示的页码
        right_has_more = False
        #判断当前页是不是比4小，比如当前页是第二页，他就不能存在 0.1.2.3.4这种情况。
        if current_page <= around_count + 2:
            left_page = range(1, current_page)
        else:
            left_has_more = True
            left_page = range(current_page-around_count, current_page)
        if current_page >= num_page-around_count-1:
            right_page = range(current_page+1, num_page+1)
        else:
            right_has_more = True
            right_page = range(current_page+1, current_page+3)
        return {
            'left_pages': left_page,
            'right_pages': right_page,
            'current_page': current_page,
            'left_has_more': left_has_more,
            'right_has_more': right_has_more,
            'finding_content':finding_content,
            'audit_department':audit_department,
        }

#级联选择IATF_IATF_Deatil
def ajax_load_finding(request):
    audit_id=request.GET.get('audit_id',None)
    if audit_id is not None:
        finding = is_finding.objects.filter(audit_id=audit_id).order_by('finding_id')
    else:
        finding = ''
    return render(request,'QtCompetence/finding_dropdown_list_options.html',{'finding':finding})

#is_action input
class action_info_create(CreateView):
    template_name = 'QtCompetence/create_action_with_dropdownlist.html'
    #model = is_action
    success_url = reverse_lazy('QtCompetence:action_info_list')
    #fields = ['responsible', 'department', 'due_date', 'actual_date', 'finding','action_content']
    form_class=actionCreateForm

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        real_name = self.request.session.get('real_name', None)
        user_pk = self.request.session.get('user_pk', None)
        is_staff = self.request.session.get('is_staff',None)
        is_manager = self.request.session.get('is_manager',None)
        is_departmentManager = self.request.session.get('is_departmentManager',None)
        is_plantManager = self.request.session.get('is_plantManager',None)
        is_group_leader = self.request.session.get('is_group_leader',None)
        context['real_name'] = real_name
        context['is_staff']=is_staff
        context['user_pk']=user_pk
        context['is_manager']=is_manager
        context['is_departmentManager']=is_department
        context['is_plantManager']=is_plantManager
        context['is_group_leader']=is_group_leader

        return context

    def post(self, request, *args, **kwargs):
        user_pk = self.request.session.get('user_pk', None)
        if user_pk:
            due_date = request.POST.get("due_date", None)
            actual_date = request.POST.get("actual_date", None)
            #这里createview中有Bug forms中设置可为非必填，数据库中可以插入空记录，但是利用create方法进行新建时候，强制DatetimeField不可以为空。
            if actual_date is '' and due_date is '':
                actual_date=datetime.datetime.now()
                due_date=datetime.datetime.now()
                action_content = request.POST.get("action_content", None)
                finding = request.POST.get("finding", None)
                rootCause = request.POST.get("rootCause", None)
                correction = request.POST.get("correction", None)
                user = request.POST.get("user", None)
                is_action.objects.create(due_date=due_date,actual_date=actual_date,finding_id=finding,
                                    action_content=action_content,rootCause=rootCause,correction=correction,user_id=user)
            elif actual_date is '' and due_date is not '':
                actual_date=datetime.datetime.now()
                action_content = request.POST.get("action_content", None)
                finding = request.POST.get("finding", None)
                rootCause = request.POST.get("rootCause", None)
                correction = request.POST.get("correction", None)
                user = request.POST.get("user", None)
                is_action.objects.create(due_date=due_date,actual_date=actual_date,finding_id=finding,
                                        action_content=action_content,rootCause=rootCause,correction=correction,user_id=user)
            elif actual_date is not '' and due_date is '':
                due_date=datetime.datetime.now()
                action_content = request.POST.get("action_content", None)
                finding = request.POST.get("finding", None)
                rootCause = request.POST.get("rootCause", None)
                correction = request.POST.get("correction", None)
                user = request.POST.get("user", None)
                is_action.objects.create(due_date=due_date,actual_date=actual_date,finding_id=finding,
                                        action_content=action_content,rootCause=rootCause,correction=correction,user_id=user)
            else:
                action_content = request.POST.get("action_content", None)
                finding = request.POST.get("finding", None)
                rootCause = request.POST.get("rootCause", None)
                correction = request.POST.get("correction", None)
                user = request.POST.get("user", None)
                is_action.objects.create(due_date=due_date,actual_date=actual_date,finding_id=finding,
                                        action_content=action_content,rootCause=rootCause,correction=correction,user_id=user)
            return redirect('QtCompetence:action_info_list')
        else:
            return redirect('QtCompetence:login_user')

#is_action update
class action_info_update(UpdateView):
    template_name = 'QtCompetence/update_action_with_dropdownlist.html'
    #model = is_action
    success_url = reverse_lazy('QtCompetence:action_info_list')
    #fields = ['responsible', 'department', 'due_date', 'actual_date', 'finding','action_content']
    form_class=actionCreateForm
    queryset=is_action.objects.all().order_by('action_id')

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        real_name = self.request.session.get('real_name', None)
        user_pk = self.request.session.get('user_pk', None)
        is_staff = self.request.session.get('is_staff',None)
        is_manager = self.request.session.get('is_manager',None)
        is_departmentManager = self.request.session.get('is_departmentManager',None)
        is_plantManager = self.request.session.get('is_plantManager',None)
        is_group_leader = self.request.session.get('is_group_leader',None)
        context['real_name'] = real_name
        context['is_staff']=is_staff
        context['user_pk']=user_pk
        context['is_manager']=is_manager
        context['is_departmentManager']=is_department
        context['is_plantManager']=is_plantManager
        context['is_group_leader']=is_group_leader
        context['return_type']='action_info'
        return context

    def post(self, request, *args, **kwargs):
        user_pk = self.request.session.get('user_pk', None)
        if user_pk:
            due_date = request.POST.get("due_date", None)
            actual_date = request.POST.get("actual_date", None)
            action_content = request.POST.get("action_content", None)
            finding = request.POST.get("finding", None)
            user = request.POST.get("user", None)
            rootCause = request.POST.get("rootCause", None)
            correction = request.POST.get("correction", None)
            user_pk = self.request.session.get('user_pk', None)
            is_action.objects.filter(action_id=self.kwargs['pk']).update(due_date=due_date,actual_date=actual_date,create_time=datetime.datetime.now(),
                                    action_content=action_content,rootCause=rootCause,correction=correction,user_id=user)
            return redirect('QtCompetence:action_info_list')
        else:
            return redirect('QtCompetence:login_user')

#is_action delete
class action_info_delete(DeleteView):
    template_name = 'QtCompetence/delete.html'
    model = is_action
    success_url = reverse_lazy('QtCompetence:action_info_list')

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        real_name = self.request.session.get('real_name', None)
        user_pk = self.request.session.get('user_pk', None)
        is_staff = self.request.session.get('is_staff',None)
        is_manager = self.request.session.get('is_manager',None)
        is_departmentManager = self.request.session.get('is_departmentManager',None)
        is_plantManager = self.request.session.get('is_plantManager',None)
        is_group_leader = self.request.session.get('is_group_leader',None)
        context['real_name'] = real_name
        context['is_staff']=is_staff
        context['user_pk']=user_pk
        context['is_manager']=is_manager
        context['is_departmentManager']=is_department
        context['is_plantManager']=is_plantManager
        context['is_group_leader']=is_group_leader
        context['return_type']='action_info'
        return context

    def post(self, request, *args, **kwargs):
        user_pk = self.request.session.get('user_pk', None)
        if user_pk:
            user_pk = self.request.session.get('user_pk', None)
            is_action.objects.filter(action_id=self.kwargs['pk']).delete()
            return redirect('QtCompetence:action_info_list')
        else:
            return redirect('QtCompetence:login_user')

#is_action list
class action_info_list(ListView):
    #model = is_action
    template_name = 'QtCompetence/action_info_list.html'
    #queryset = is_action.objects.all().order_by("-create_time")
    paginate_by = 20#一个页面显示的条目
    context_object_name = 'action_info'


    def get_queryset(self):
        return action_info_filter(self.request.GET,queryset=is_action.objects.all().order_by("-finding__audit__audit_date")).qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        real_name = self.request.session.get('real_name', None)
        user_pk = self.request.session.get('user_pk', None)
        is_staff = self.request.session.get('is_staff',None)
        is_manager = self.request.session.get('is_manager',None)
        is_departmentManager = self.request.session.get('is_departmentManager',None)
        is_plantManager = self.request.session.get('is_plantManager',None)
        action_filter = action_info_filter(self.request.GET,queryset=is_action.objects.all().order_by("-finding__audit__audit_date"))
        is_group_leader = self.request.session.get('is_group_leader',None)
        context['real_name'] = real_name
        context['is_staff']=is_staff
        context['user_pk']=user_pk
        context['is_manager']=is_manager
        context['is_departmentManager']=is_department
        context['is_plantManager']=is_plantManager
        context['is_group_leader']=is_group_leader
        context['action_filter'] = action_filter
        paginator = context.get('paginator')
        page_obj = context.get('page_obj')
        user = self.request.GET.get('user','')
        finding = self.request.GET.get('finding','')
        paginator_data = self.get_pagination_data(paginator, page_obj,user,finding)
        context.update(paginator_data) # 将当前字典的kv更新到context字典中。
        return context

    # 这里来负责跳转的页码处理
    def get_pagination_data(self, paginator, page_obj, user='', finding='', around_count=2): # arount_count=2表示从当前页前推两页，后推两页
        current_page = page_obj.number
        num_page = paginator.num_pages
        left_has_more = False # 左边还有没有未显示的页码
        right_has_more = False
        #判断当前页是不是比4小，比如当前页是第二页，他就不能存在 0.1.2.3.4这种情况。
        if current_page <= around_count + 2:
            left_page = range(1, current_page)
        else:
            left_has_more = True
            left_page = range(current_page-around_count, current_page)
        if current_page >= num_page-around_count-1:
            right_page = range(current_page+1, num_page+1)
        else:
            right_has_more = True
            right_page = range(current_page+1, current_page+3)
        return {
            'left_pages': left_page,
            'right_pages': right_page,
            'current_page': current_page,
            'left_has_more': left_has_more,
            'right_has_more': right_has_more,
            'user':user,
            'finding':finding,
        }

# #excel导出数据处理
# import os,sys
# from public.utils import *

#转化原生sql为dict字典
def dictfetchall(cursor):
        "Return all rows from a cursor as a dict"
        columns = [col[0] for col in cursor.description]
        return [
            dict(zip(columns, row))
            for row in cursor.fetchall()
        ]
# query export page
#审核日期	迎审部门	审核主体	审核类型	审核类型信息	线体/产品	审核负责人/迎审人	创建日期	问题内容	IATF条款（带描述）	IATF详细条款（带描述）	问题类型	问题责任部门	创建日期	根本原因	纠正	纠正措施	纠正措施截止日期	实际纠正措施日期

#筛选audit数据的sql
def audit_all_info_export_sql(cursor):
    cursor.execute('''select d.audit_id as audit_id,
                        d.audit_date as audit_date,
                        d.audit_department_name as audit_department_name,
                        d.audit_body_name as audit_body_name,
                        d.audit_body_detail_name as audit_body_detail_name,
                        d.audit_scope as audit_scope,
                        d.audit_type_name as audit_type_name,
                        d.line as line,
                        d.auditor as auditor,
                        d.audit_create_time as audit_create_time,
                        d.finding_content as finding_content,
                        d.IATF_title_ch as IATF_title_ch,
                        d.IATF_detail_content_ch as IATF_detail_content_ch,
                        d.weight,
                        d.finding_audit_department as finding_audit_department,
                        d.auditee as auditee,
                        d.finding_create_time as finding_create_time,
                        e.rootCause as rootCause,
                        e.correction as correction,
                        e.action_content as action_content,
                        e.due_date as action_due_date,
                        e.actual_date as action_actual_date,
                        (select c.employ_name from iq.qtcompetence_is_user as c where c.id=e.user_id)as action_responsible,
                        e.create_time as action_create_time
                        from (select a.audit_id as audit_id,
                        a.audit_date as audit_date,
                        (select c.audit_department_name from iq.qtcompetence_is_audit_department as c where c.audit_department_id=a.department_id)as audit_department_name,
                        (select c.audit_body_name from iq.qtcompetence_is_audit_body as c where c.audit_body_id=a.audit_body_id)as audit_body_name,
                        (select c.audit_body_detail_name from iq.qtcompetence_is_audit_body_detail as c where c.audit_body_detail_id=a.audit_body_detail_id)as audit_body_detail_name,
                        a.audit_scope as audit_scope,
                        (select c.audit_type_name from iq.qtcompetence_is_audit_type as c where c.audit_type_id=a.audit_type_id)as audit_type_name,
                        a.line as line,
                        (select c.employ_name from iq.qtcompetence_is_user as c where c.id=a.user_id)as auditor,
                        a.create_time as audit_create_time,
                        b.finding_content as finding_content,
                        (select c.IATF_title_ch from iq.qtcompetence_is_IATF as c where c.IATF_id=b.IATF_id)as IATF_title_ch,
                        (select c.IATF_detail_content_ch from iq.qtcompetence_is_IATF_detail as c where c.IATF_detail_id=b.IATF_detail_id)as IATF_detail_content_ch,
                        b.weight as weight,
                        (select c.audit_department_name from iq.qtcompetence_is_audit_department as c where c.audit_department_id=b.audit_department_id)as finding_audit_department,
                        (select c.employ_name from iq.qtcompetence_is_user as c where c.id=b.user_id)as auditee,
                        b.create_time as finding_create_time,
                        b.finding_id as finding_id from 
                        iq.qtcompetence_is_audit as a left join iq.qtcompetence_is_finding as b on b.audit_id = a.audit_id) as d left join iq.qtcompetence_is_action as e on e.finding_id=d.finding_id''');
    return dictfetchall(cursor)

#筛选competence数据的sql
def competence_all_info_export_sql(cursor):
    cursor.execute('''select  (select b.competence_name from iq.qtcompetence_is_competition_rule as b where a.competence_id=b.competence_id) as competence_name,
                    (select b.employ_name from iq.qtcompetence_is_user as b where b.id=a.user_id) as user_name,
                    (select b.user_name from iq.qtcompetence_is_user as b where b.id=a.user_id) as badge_number,
                    (select c.position_name from iq.qtcompetence_is_position as c where c.id= (select b.position_id from iq.qtcompetence_is_user as b where b.id=a.user_id)) as position_name,
                    (select d.standard_score-12 from iq.qtcompetence_r_rules_position as d where d.position_id= (select e.position_id from iq.qtcompetence_is_user as e where a.user_id=e.id) and a.competence_id=d.rules_id) as standard_score,
                    self_assessment-12 as self_assessment ,self_gap ,manager_assessment-12 as manager_assessment,manager_gap, date_format(create_time,"%Y-%m-%d") as time from iq.qtcompetence_is_competition_score as a order by a.assessment_id;''');
    return dictfetchall(cursor)

#audit all infomation list
class audit_all_info_list(ListView):
 #model = is_finding
    template_name = 'QtCompetence/audit_all_info_list.html'
    paginate_by = 30#一个页面显示的条目
    context_object_name = 'audit_info'

    def get_queryset(self):
        with connection.cursor() as cursor:
            all_info=audit_all_info_export_sql(cursor)
        return all_info
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        real_name = self.request.session.get('real_name', None)
        user_pk = self.request.session.get('user_pk', None)
        is_staff = self.request.session.get('is_staff',None)
        is_manager = self.request.session.get('is_manager',None)
        is_departmentManager = self.request.session.get('is_departmentManager',None)
        is_plantManager = self.request.session.get('is_plantManager',None)
        is_group_leader = self.request.session.get('is_group_leader',None)
        context['real_name'] = real_name
        context['is_staff']=is_staff
        context['user_pk']=user_pk
        context['is_manager']=is_manager
        context['is_departmentManager']=is_department
        context['is_plantManager']=is_plantManager
        context['is_group_leader']=is_group_leader
        paginator = context.get('paginator')
        page_obj = context.get('page_obj')
        paginator_data = self.get_pagination_data(paginator, page_obj)
        context.update(paginator_data) # 将当前字典的kv更新到context字典中。
        return context

    # 这里来负责跳转的页码处理
    def get_pagination_data(self, paginator, page_obj,around_count=2): # arount_count=2表示从当前页前推两页，后推两页
        current_page = page_obj.number
        num_page = paginator.num_pages
        left_has_more = False # 左边还有没有未显示的页码
        right_has_more = False
        #判断当前页是不是比4小，比如当前页是第二页，他就不能存在 0.1.2.3.4这种情况。
        if current_page <= around_count + 2:
            left_page = range(1, current_page)
        else:
            left_has_more = True
            left_page = range(current_page-around_count, current_page)
        if current_page >= num_page-around_count-1:
            right_page = range(current_page+1, num_page+1)
        else:
            right_has_more = True
            right_page = range(current_page+1, current_page+3)
        return {
            'left_pages': left_page,
            'right_pages': right_page,
            'current_page': current_page,
            'left_has_more': left_has_more,
            'right_has_more': right_has_more,
        }

#competence all infomation list
class competence_all_info_list(ListView):
    template_name = 'QtCompetence/competence_all_info_list.html'
    paginate_by = 30#一个页面显示的条目
    context_object_name = 'competence_info'

    def get_queryset(self):
        with connection.cursor() as cursor:
            all_info=competence_all_info_export_sql(cursor)
        return all_info
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        real_name = self.request.session.get('real_name', None)
        user_pk = self.request.session.get('user_pk', None)
        is_staff = self.request.session.get('is_staff',None)
        is_manager = self.request.session.get('is_manager',None)
        is_departmentManager = self.request.session.get('is_departmentManager',None)
        is_plantManager = self.request.session.get('is_plantManager',None)
        is_group_leader = self.request.session.get('is_group_leader',None)
        context['real_name'] = real_name
        context['is_staff']=is_staff
        context['user_pk']=user_pk
        context['is_manager']=is_manager
        context['is_departmentManager']=is_department
        context['is_plantManager']=is_plantManager
        context['is_group_leader']=is_group_leader
        paginator = context.get('paginator')
        page_obj = context.get('page_obj')
        paginator_data = self.get_pagination_data(paginator, page_obj)
        context.update(paginator_data) # 将当前字典的kv更新到context字典中。
        return context

    # 这里来负责跳转的页码处理
    def get_pagination_data(self, paginator, page_obj,around_count=2): # arount_count=2表示从当前页前推两页，后推两页
        current_page = page_obj.number
        num_page = paginator.num_pages
        left_has_more = False # 左边还有没有未显示的页码
        right_has_more = False
        #判断当前页是不是比4小，比如当前页是第二页，他就不能存在 0.1.2.3.4这种情况。
        if current_page <= around_count + 2:
            left_page = range(1, current_page)
        else:
            left_has_more = True
            left_page = range(current_page-around_count, current_page)
        if current_page >= num_page-around_count-1:
            right_page = range(current_page+1, num_page+1)
        else:
            right_has_more = True
            right_page = range(current_page+1, current_page+3)
        return {
            'left_pages': left_page,
            'right_pages': right_page,
            'current_page': current_page,
            'left_has_more': left_has_more,
            'right_has_more': right_has_more,
        }

#switch audit_scope
switch_audit_scope = lambda x: {
    x == 16: 'internal', x == 17: 'external',x != 16 and x != 17:'None'
}

#switch finding_weight
switch_finding_weight = lambda x: {
    x == 18: 'Major Nonconfirming', x == 19: 'Minor Nonconfirming',x == 20: 'OFI',x != 18 and x != 19 and x != 20:'None'
}

#audit excel export
def export_audit_all_info_excel(request):
    wb = Workbook()		# 生成一个工作簿（即一个Excel文件）
    wb.encoding = 'utf-8'
    sheet1 = wb.active	# 获取第一个工作表（sheet1）
    sheet1.title = 'Audit_all_info'	# 给工作表1设置标题
    '''Audit_id 审核日期	迎审部门	审核主体  审核主体详情	审核类型	审核类型信息	线体/产品	审核负责人/迎审人  Audit创建日期	
                问题内容	IATF条款（带描述）	IATF详细条款（带描述）	问题类型	问题责任部门	问题负责人   Finding创建日期
                根本原因	纠正	纠正措施	纠正措施截止日期	实际纠正措施日期  纠正负责人  Action创建日期
    row_one = ['Audit_id','审核日期', '迎审部门', '审核主体', '审核主体详情', '审核类型', '审核类型信息', '线体/产品', '审核负责人/迎审人', 'Audit创建日期',
               '问题内容', 'IATF条款（带描述）', 'IATF详细条款（带描述）', '问题类型', '问题责任部门', '问题负责人', 'Finding创建日期',
               '根本原因', '纠正', '纠正措施', '纠正措施截止日期', '实际纠正措施日期', '纠正负责人', 'Action创建日期']'''
    row_one = ['Audit_id','审核日期', '迎审部门', '审核主体', '审核主体详情', '审核类型', '审核类型信息', '线体/产品', '审核负责人/迎审人', 'Audit创建日期',
               '问题内容', 'IATF条款（带描述）', 'IATF详细条款（带描述）', '问题类型', '问题责任部门', '问题负责人', 'Finding创建日期',
               '根本原因', '纠正', '纠正措施', '纠正措施截止日期', '实际纠正措施日期', '纠正负责人', 'Action创建日期']
    for i in range(1, len(row_one)+1):	# 从第一行开始写，因为Excel文件的行号是从1开始，列号也是从1开始
        # 从row=1，column=1开始写，即将row_one的数据依次写入第一行
        sheet1.cell(row=1, column=i).value=row_one[i-1]
    with connection.cursor() as cursor:
        all_info=audit_all_info_export_sql(cursor)
        if  all_info is not None:
            for audit in all_info:
                max_row = sheet1.max_row + 1	# 获取到工作表的最大行数并加1
                audit_info = [audit['audit_id'],str(audit['audit_date'])[:10],audit['audit_department_name'],audit['audit_body_name'],audit['audit_body_detail_name'],switch_audit_scope(audit['audit_scope'])[True],audit['audit_type_name'],audit['line'],audit['auditor'],str(audit['audit_create_time'])[:10],
                              audit['finding_content'],audit['IATF_title_ch'],audit['IATF_detail_content_ch'],switch_finding_weight(audit['weight'])[True],audit['finding_audit_department'],audit['auditee'],str(audit['finding_create_time'])[:10],
                              audit['rootCause'],audit['correction'],audit['action_content'],str(audit['action_actual_date'])[:10],str(audit['action_responsible'])[:10],audit['action_responsible'],str(audit['action_create_time'])[:10]]
                for x in range(1, len(audit_info)+1):		# 将每一个对象的所有字段的信息写入一行内
                    sheet1.cell(row=max_row, column=x).value = audit_info[x-1]
        else:
            pass
    # 准备写入到IO中
    output = BytesIO()
    wb.save(output)	 # 将Excel文件内容保存到IO中
    output.seek(0)	 # 重新定位到开始
    # 设置HttpResponse的类型
    response = HttpResponse(output.getvalue(), content_type='application/vnd.ms-excel')
    ctime = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    file_name = 'Audit_all_info%s.xlsx' % ctime	# 给文件名中添加日期时间
    file_name = urlquote(file_name)	 # 使用urlquote()方法解决中文无法使用的问题
    response['Content-Disposition'] = 'attachment; filename=%s' % file_name
    # response.write(output.getvalue())	 # 在设置HttpResponse的类型时，如果给了值，可以不写这句
    return response

#competence excel export
def export_competence_all_info_excel(request):
    wb = Workbook()		# 生成一个工作簿（即一个Excel文件）
    wb.encoding = 'utf-8'
    sheet1 = wb.active	# 获取第一个工作表（sheet1）
    sheet1.title = 'competence_all_info'	# 给工作表1设置标题
    row_one = ['能力测评项','姓名', '员工编号', '职位', '标准分数', '自评分数', '自评差距', '经理评价分数', '经理评价差距', '日期']
    for i in range(1, len(row_one)+1):	# 从第一行开始写，因为Excel文件的行号是从1开始，列号也是从1开始
        # 从row=1，column=1开始写，即将row_one的数据依次写入第一行
        sheet1.cell(row=1, column=i).value=row_one[i-1]
    with connection.cursor() as cursor:
        all_info=competence_all_info_export_sql(cursor)
        if  all_info is not None:
            for competence in all_info:
                max_row = sheet1.max_row + 1	# 获取到工作表的最大行数并加1
                competence_info = [competence['competence_name'],competence['user_name'],competence['badge_number'],competence['position_name'],competence['standard_score'],
                                competence['self_assessment'],competence['self_gap'],competence['manager_assessment'],competence['manager_gap'],competence['time']]
                for x in range(1, len(competence_info)+1):		# 将每一个对象的所有字段的信息写入一行内
                    sheet1.cell(row=max_row, column=x).value = competence_info[x-1]
        else:
            pass
    # 准备写入到IO中
    output = BytesIO()
    wb.save(output)	 # 将Excel文件内容保存到IO中
    output.seek(0)	 # 重新定位到开始
    # 设置HttpResponse的类型
    response = HttpResponse(output.getvalue(), content_type='application/vnd.ms-excel')
    ctime = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    file_name = 'Competence_info%s.xlsx' % ctime	# 给文件名中添加日期时间
    file_name = urlquote(file_name)	 # 使用urlquote()方法解决中文无法使用的问题
    response['Content-Disposition'] = 'attachment; filename=%s' % file_name
    # response.write(output.getvalue())	 # 在设置HttpResponse的类型时，如果给了值，可以不写这句
    return response

#对接KnowHow search功能
def api_search(request):
    feedbackInfo=[]
    if request.method=='GET':
        search = request.GET.get('search',default="8.5")
        #is_audit
        if is_user.objects.filter(employ_name__icontains=search).exists():
            #有详情页面这么写
            # for i in range(0,len(is_user.objects.filter(employ_name__icontains=search))-1):
            #     if is_audit.objects.filter(user_id=is_user.objects.filter(employ_name__icontains=search)[i].id).exists():
            #         feedbackInfo.append({"System":"IQ_Audit","link":"http://10.192.104.156/audit_info_list?department=&user="+search+"&audit_body=","Title":search+"的审核要求信息","Abstract":str(is_audit.objects.filter(user_id=is_user.objects.filter(employ_name__icontains=search)[i].id)[0].line)+"线的审核要求信息","Created":str(is_audit.objects.filter(user_id=(is_user.objects.filter(employ_name__icontains=search)[i].id))[0].create_time.strftime("%Y-%m-%d, %H:%M:%S"))})
            #     if is_action.objects.filter(user_id=is_user.objects.filter(employ_name__icontains=search)[i].id).exists():
            #         feedbackInfo.append({"System":"IQ_Audit","link":"http://10.192.104.156/action_info_list?user="+search+"&finding=","Title":search+"的fingding改进信息详情","Abstract":str(is_action.objects.filter(user_id=is_user.objects.filter(employ_name__icontains=search)[i].id)[0].action_content),"Created":str(is_action.objects.filter(user_id=(is_user.objects.filter(employ_name__icontains=search)[i].id))[0].create_time.strftime("%Y-%m-%d, %H:%M:%S"))})
            #只显示第一条记录
            for i in range(0,len(is_user.objects.filter(employ_name__icontains=search))-1):
                if is_audit.objects.filter(user_id=is_user.objects.filter(employ_name__icontains=search)[i].id).exists():
                    feedbackInfo.append({"System":"IQ_Audit","link":"http://10.192.104.156/audit_info_list?department=&user="+search+"&audit_body=","Abstract":"关于"+search+"的审核要求信息详情","Created":str(is_audit.objects.filter(user_id=is_user.objects.filter(employ_name__icontains=search)[i].id)[0].create_time.strftime("%Y-%m-%d %H:%M:%S"))})
                    # feedbackInfo.append(',')
                    break
            for i in range(0,len(is_user.objects.filter(employ_name__icontains=search))-1):
                if is_action.objects.filter(user_id=is_user.objects.filter(employ_name__icontains=search)[i].id).exists():
                    feedbackInfo.append({"System":"IQ_Audit","link":"http://10.192.104.156/action_info_list?user="+search+"&finding=","Abstract":"关于"+search+"的fingding改进信息详情","Created":str(is_action.objects.filter(user_id=is_user.objects.filter(employ_name__icontains=search)[i].id)[0].create_time.strftime("%Y-%m-%d %H:%M:%S"))})
                    # feedbackInfo.append(',')
                    break
        else:
            # if is_audit_department.objects.filter(audit_department_name=search).exists:
            #     if is_audit.objects.filter(department_id=is_audit_department.objects.filter(audit_department_name=search)[0].audit_department_id).exists():
            #         feedbackInfo.append({"System":"IQ_Audit","link":"http://10.192.104.156/audit_info_list?department="+is_audit_department.objects.filter(audit_department_name=search)[0].audit_department_id+"&user=&audit_body=","Title":search+"的审核要求信息详情","Abstract":is_audit.objects.filter(department_id=is_audit_department.objects.filter(audit_department_name=search)[0].audit_department_id)[0].line+"线的审核要求信息","Created":str(is_audit.objects.filter(department_id=is_audit_department.objects.filter(audit_department_name=search)[0].audit_department_id)[0].create_time.strftime("%Y-%m-%d, %H:%M:%S"))+","})
            # if is_audit_body.objects.filter(audit_body_name=search).exists():
            #     if is_audit.objects.filter(audit_body_id=is_audit_body.objects.filter(audit_body_name=search)[0].audit_body_id).exists():
            #         feedbackInfo.append({"System":"IQ_Audit","link":"http://10.192.104.156/audit_info_list?department=&user=&audit_body="+is_audit_body.objects.filter(audit_body_name=search)[0].audit_body_id+"","Title":search+"的审核要求信息详情","Abstract":is_audit.objects.filter(audit_body_id=is_audit_body.objects.filter(audit_body_name=search)[0].audit_body_id)[0].line+"线的审核要求信息","Created":str(is_audit.objects.filter(audit_body_id=is_audit_body.objects.filter(audit_body_name=search)[0].audit_body_id)[0].create_time.strftime("%Y-%m-%d, %H:%M:%S"))+","})
            if is_audit_department.objects.filter(audit_department_name=search).exists:
                for i in range(0,len(is_audit_department.objects.filter(audit_department_name=search))-1):
                    if is_audit.objects.filter(department_id=is_audit_department.objects.filter(audit_department_name=search)[i].audit_department_id).exists():
                        feedbackInfo.append({"System":"IQ_Audit","link":"http://10.192.104.156/audit_info_list?department="+is_audit_department.objects.filter(audit_department_name=search)[i].audit_department_id+"&user=&audit_body=","Title":search+"的审核要求信息详情","Abstract":is_audit.objects.filter(department_id=is_audit_department.objects.filter(audit_department_name=search)[i].audit_department_id)[0].line+"线的审核要求信息","Created":str(is_audit.objects.filter(department_id=is_audit_department.objects.filter(audit_department_name=search)[i].audit_department_id)[0].create_time.strftime("%Y-%m-%d %H:%M:%S"))})
                        # feedbackInfo.append(',')
                        break
            if is_audit_body.objects.filter(audit_body_name=search).exists():
                for i in range(0,len(is_audit_body.objects.filter(audit_body_name=search))-1):
                    if is_audit.objects.filter(audit_body_id=is_audit_body.objects.filter(audit_body_name=search)[i].audit_body_id).exists():
                        feedbackInfo.append({"System":"IQ_Audit","link":"http://10.192.104.156/audit_info_list?department=&user=&audit_body="+is_audit_body.objects.filter(audit_body_name=search)[i].audit_body_id+"","Title":search+"的审核要求信息详情","Abstract":is_audit.objects.filter(audit_body_id=is_audit_body.objects.filter(audit_body_name=search)[0].audit_body_id)[0].line+"线的审核要求信息","Created":str(is_audit.objects.filter(audit_body_id=is_audit_body.objects.filter(audit_body_name=search)[i].audit_body_id)[0].create_time.strftime("%Y-%m-%d %H:%M:%S"))})
                        # feedbackInfo.append(',')
                        break
        #is_finding
        if is_finding.objects.filter(finding_content__icontains=search).exists():
            feedbackInfo.append({"System":"IQ_Audit","link":"http://10.192.104.156/finding_info_list?finding_content="+search+"&audit_department=&audit=","Title":search+"的finding信息详情","Abstract":is_finding.objects.filter(finding_content__icontains=search)[0].finding_content,"Created":str(is_finding.objects.filter(finding_content__icontains=search)[0].create_time.strftime("%Y-%m-%d %H:%M:%S"))})
            # feedbackInfo.append(',')        
        if is_audit_department.objects.filter(audit_department_name=search).exists():
            if is_finding.objects.filter(audit_department_id=is_audit_department.objects.filter(audit_department_name=search)[0].audit_department_id).exists():
                feedbackInfo.append({"System":"IQ_Audit","link":"http://10.192.104.156/audit_info_list?department="+str(is_audit_department.objects.filter(audit_department_name=search)[0].audit_department_id)+"&user=&audit_body=","Title":search+"的审核要求信息详情","Abstract":str(is_finding.objects.filter(audit_department_id=is_audit_department.objects.filter(audit_department_name=search)[0].audit_department_id)[0].finding_content),"Created":str(is_finding.objects.filter(audit_department_id=is_audit_department.objects.filter(audit_department_name=search)[0].audit_department_id)[0].create_time.strftime("%Y-%m-%d %H:%M:%S"))})
                # feedbackInfo.append(',')
        #is_action
        # if is_user.objects.filter(employ_name__icontains=search).exists():
        #     if is_audit.objects.filter(user_id=is_user.objects.filter(employ_name__icontains=search)[0].id).exists():
        #         feedbackInfo.append({"System":"IQ_Audit","link":"http://10.192.104.156/audit_info_list?department=&user="+search+"&audit_body=","Abstract":"关于"+search+"的审核要求信息详情","Created":is_audit.objects.filter(user_id=is_user.objects.filter(employ_name__icontains=search)[0].id)[0].create_time})
        #     elif is_finding.objects.filter(user_id=is_user.objects.filter(employ_name__icontains=search)[0].id).exists():
        #         feedbackInfo.append({"System":"IQ_Audit","link":"http://10.192.104.156/finding_info_list","Abstract":"关于"+search+"的finding信息详情","Created":is_audit.objects.filter(line__icontains=search)[0].create_time})
        #     elif is_action.objects.filter(user_id=is_user.objects.filter(employ_name__icontains=search)[0].id).exists():
        #         feedbackInfo.append({"System":"IQ_Audit","link":"http://10.192.104.156/action_info_list","Abstract":"关于"+search+"的fingding改进信息详情","Created":is_audit.objects.filter(line__icontains=search)[0].create_time})
        # else:
        #     if is_audit_department.objects.filter(audit_department_name=search).exists:
        #         if is_audit.objects.filter(department_id=is_audit_department.objects.filter(audit_department_name=search)[0].audit_department_id).exists():
        #             feedbackInfo.append({"System":"IQ_Audit","link":"http://10.192.104.156/audit_info_list?department="+is_audit_department.objects.filter(audit_department_name=search)[0].audit_department_id+"&user=&audit_body=","Abstract":"关于"+search+"的审核要求信息详情","Created":is_audit.objects.filter(department_id=is_audit_department.objects.filter(audit_department_name=search)[0].audit_department_id)[0].create_time})
        #     if is_audit_body.objects.filter(audit_body_name=search).exists():
        #         if is_audit.objects.filter(audit_body_id=is_audit_body.objects.filter(audit_body_name=search)[0].audit_body_id).exists():
        #             feedbackInfo.append({"System":"IQ_Audit","link":"http://10.192.104.156/audit_info_list?department=&user=&audit_body="+is_audit_body.objects.filter(audit_body_name=search)[0].audit_body_id+"","Abstract":"关于"+search+"的审核要求信息详情","Created":is_audit.objects.filter(audit_body_id=is_audit_body.objects.filter(audit_body_name=search)[0].audit_body_id)[0].create_time})
    return HttpResponse(json.dumps(feedbackInfo))

# 引入发送邮件的模块
from django.core.mail import send_mail, send_mass_mail, EmailMultiAlternatives
def send_email(request,user_id):
    address="feng.an@continental-corporation.com"
    res=send_mail('放假通知','中秋放假三天',address,[address])
    if res==1:
        return HttpResponse('success')
    else:
        return HttpResponse('fail')

def contact_us(request):
    return HttpResponse('contact_us')

def FAQ(request):
    return HttpResponse('FAQ')

def admin(request):
    return redirect('admin:login')

#封装的数据库语句执行方法
def exec_sql(sql, params=None, db='default'):
    """
    执行sql，例如insert和update
    :param sql: sql语句
    :param params: sql语句参数
    :param db: Django数据库名
    """
    cursor = connections[db].cursor()
    cursor.execute(sql, params)
    cursor.close()
    return True


def fetchone_sql(sql, params=None, db='default', flat=False):
    """
    返回一行数据
    :param sql: sql语句
    :param params: sql语句参数
    :param db: Django数据库名
    :param flat: 如果为True，只返回第一个字段值，例如：id
    :return: 例如：(id, 'username', 'first_name')
    """
    cursor = connections[db].cursor()
    cursor.execute(sql, params)
    fetchone = cursor.fetchone()
    cursor.close()
    if fetchone:
        fetchone = fetchone[0] if flat else fetchone
    return fetchone


def fetchone_to_dict(sql, params=None, db='default'):
    """
    返回一行数据
    :param sql: sql语句
    :param params: sql语句参数
    :param db: Django数据库名
    :return: 例如：{"id": id, "username": 'username', "first_name": 'first_name'}
    """
    cursor = connections[db].cursor()
    cursor.execute(sql, params)
    desc = cursor.description
    row = dict(zip([col[0] for col in desc], cursor.fetchone()))
    cursor.close()
    return row


def fetchall_sql(sql, params=None, db='default', flat=False):
    """
    返回全部数据
    :param sql: sql语句
    :param params: sql语句参数
    :param db: Django数据库名
    :param flat: 如果为True，只返回每行数据第一个字段值的元组，例如：(id1, id2, id3)
    :return: 例如：[(id, 'username', 'first_name')]
    """
    cursor = connections[db].cursor()
    cursor.execute(sql, params)
    fetchall = cursor.fetchall()
    cursor.close()
    if fetchall:
        fetchall = tuple([o[0] for o in fetchall]) if flat else fetchall
    return fetchall


def fetchall_to_dict(sql, params=None, db='default'):
    """
    返回全部数据
    :param sql: sql语句
    :param params: sql语句参数
    :param db: Django数据库名
    :return: 例如：[{"id": id, "username": 'username', "first_name": 'first_name'}]
    """
    cursor = connections[db].cursor()
    cursor.execute(sql, params)
    desc = cursor.description
    object_list = [
        dict(zip([col[0] for col in desc], row))
        for row in cursor.fetchall()
    ]
    cursor.close()
    return object_list

#处理ir_dimUser中的Email_id为空情况，以及输入格式非标准化
def normallizeEmail(email):
    if email is None:
        return None
    else:
        return email.split('@')[0]
#处理ir_dimUser中的DEPARTMENTKEY为空情况，以及输入格式非标准化
#is_department is_position 第一万条记录处理空数据手动创建代表无岗位无工作
#服务端同步到本地端转换序号
def deptID_to_ID(deptID):
    if deptID is None or not is_department.objects.filter(department_id=deptID).exists():
        return 10000
    else:
        id=is_department.objects.filter(department_id=deptID)[0].id
        return id
#本地端序号转回服务端
def ID_to_deptID(ID):
    #肯定不会为空
    department_id=is_department.objects.filter(id=ID)[0].department_id
    if department_id == 10000:
        return None
    else:
        return department_id
#处理ir_dimUser中的JOB_CODE_ID为空情况，以及输入格式非标准化
#服务端同步到本地端转换序号
def posID_to_ID(posID):
    if posID is None or not is_position.objects.filter(position_id=posID).exists():
        return 10000
    else:
        id=is_position.objects.filter(position_id=posID)[0].id
        return id
#本地端序号转回服务端
def ID_to_posID(ID):
    position_id=is_position.objects.filter(id=ID)[0].position_id
    if position_id == 10000:
        return None
    else:
        return position_id
#处理level为空情况 is_user第十万条数据表示无员工初始化
#服务端同步到本地端转换序号
def userID_to_ID(userID):
    if userID is None or not is_user.objects.filter(user_id=userID).exists():
        return 100000
    else:
        id=is_user.objects.filter(user_id=userID)[0].id
        return id
#本地端序号转回服务端
def ID_to_userID(ID):
    user_id=is_user.objects.filter(id=ID)[0].user_id
    if user_id == 100000:
        return None
    else:
        return user_id
#开启定时工作
try:
    # 实例化调度器
    scheduler = BackgroundScheduler()
    # 调度器使用DjangoJobStore()
    scheduler.add_jobstore(DjangoJobStore(), "default")
    # 设置定时任务，选择方式为interval，时间间隔为10s
    # 此处为HR 远程端数据库
    dept_list=fetchall_to_dict("select DEPARTMENTKEY,DEPARTMENTNAME,MANAGER_ID,INUSE from ir_dimdepartment where INUSE='Y' order by DEPARTMENTKEY",db='hr')
    pos_list=fetchall_to_dict("select JOB_CODE_ID,JOB_TITLE,INUSE from OEE.ETF_Dim_JobCode where INUSE='Y' order by JOB_CODE_ID",db='hr')
    user_list=fetchall_to_dict("select ID,SYS_NAME,BADGE_NO,USER_NAME,EMAIL_ID,DEPARTMENTKEY,JOB_CODE_ID,LEVEL_1_R,LEVEL_2_R,LEVEL_3_R,LEVEL_4_R,INUSE from ir_dimUser where INUSE='Y'",db='hr')
    # @register_job(scheduler,"interval", seconds=600)
    # 另一种方式为每天固定时间执行任务，对应代码为：day_of_week='mon-sun'
    @register_job(scheduler, 'cron', day_of_week='mon-sun', hour='13', minute='15', second='30',id='task_time')
    def my_job():
        # 从OEE.ETF_Dim_JobCode中写入is_position中岗位信息，岗位ID等信息
        # 从ir_dimUser中写入is_user表中人员信息
        # SELECT DEPARTMENTKEY,DEPARTMENTNAME,INUSE,MANAGER_ID from ir_dimdepartment  where INUSE = 'Y' order by DEPARTMENTKEY;
        # select JOB_CODE_ID,JOB_TITLE,INUSE from OEE.ETF_Dim_JobCode where INUSE='Y' order by JOB_CODE_ID;
        # SELECT ID,SYS_NAME,BADGE_NO, USER_NAME,EMAIL_ID,DEPARTMENTKEY,JOB_CODE_ID FROM IR_DIMUSER WHERE DEPARTMENTKEY='542' AND INUSE='Y';
        # is_department
        #需要删除的数据
        need_delete_list=[]
        #需要增加的数据的列表
        need_add_list=[]
        #需要更新的数据
        need_update_list=[]
        qtDept=is_department.objects.all()
        if not qtDept:
            #首次插入为初始化插入
                for i in range(0,len(dept_list)):
                    #exec_sql('insert into iq.qtcompetence_is_department values (%(ID)s,%(DEPARTMENTKEY)s,%(DEPARTMENTNAME)s,%(TIME)s)',params={"ID":[i+1],"DEPARTMENTKEY":dept_list[i]['DEPARTMENTKEY'],"DEPARTMENTNAME":dept_list[i]['DEPARTMENTNAME'],"TIME":datetime.datetime.now()})
                    is_department.objects.create(department_id=dept_list[i]['DEPARTMENTKEY'],department_name=dept_list[i]['DEPARTMENTNAME'])
                #exec_sql('insert into iq.qtcompetence_is_department values (%(ID)s,%(DEPARTMENTKEY)s,%(DEPARTMENTNAME)s,%(TIME)s)',params={"ID":10000,"DEPARTMENTKEY":10000,"DEPARTMENTNAME":"NO DEPARTMENT","TIME":datetime.datetime.now()})
                is_department.objects.create(id=10000,department_id=10000,department_name="NO DEPARTMENT")
                print('qtcompetence_is_department is inserted at first time, time is :'+str(datetime.datetime.now()))
        else:
            #先获取目前最大索引id
            now_index=(is_department.objects.order_by('-id')[1:2])[0].id
            print(now_index)
            #1.遍历local对比远程端筛选出需要删除(本地有远程端没有)或则本地有远程端有但是INUSE='N'项进行删除
            for dept in qtDept:
                InUse=fetchall_to_dict("select INUSE from ir_dimdepartment where DEPARTMENTKEY=%(DEPARTMENTKEY)s",params={"DEPARTMENTKEY":dept.department_id},db='hr')
                if not InUse:
                    if dept.department_id != 10000:
                    #10000为手动添加No Department for django
                        need_delete_list.append(dept.id)
                else:
                    if InUse[0]['INUSE'] == 'N':
                        need_delete_list.append(dept.id)
            print("is_department's need_delete_list is:"+str(need_delete_list))
            #删除本地多余数据USER_ID
            if need_delete_list:
                for id in need_delete_list:
                    #exec_sql('delete from iq.qtcompetence_is_department where department_id=%(DEPARTMENTKEY)s',params={"DEPARTMENTKEY":id})
                    is_department.objects.filter(id= id).delete()
                print("is_department's need_delete_list is deleted, time is: "+str(datetime.datetime.now()))
            #2.遍历远程数据库对比本地筛选出需要增加(远程端有本地没有)项进行增加
            for i in range(0,len(dept_list)):
                if not (is_department.objects.filter(department_id=dept_list[i]['DEPARTMENTKEY']).exists() and dept_list[i]['INUSE'] == 'Y'):
                    need_add_list.append(dept_list[i]['DEPARTMENTKEY'])
            print("is_department's need_add_list:"+str(need_add_list))
            #直接插入本地
            if need_add_list:
                for i in range(0,len(need_add_list)):
                    hr_dept_list=fetchall_to_dict("select DEPARTMENTKEY,DEPARTMENTNAME,MANAGER_ID from ir_dimdepartment where INUSE='Y' and DEPARTMENTKEY=%(DEPARTMENTKEY)s",params={"DEPARTMENTKEY":need_add_list[i]},db='hr')
                    print(hr_dept_list)
                    is_department.objects.create(id=now_index+i+1,department_id=hr_dept_list[0]['DEPARTMENTKEY'],department_name=hr_dept_list[0]['DEPARTMENTNAME'])
            print("is_department's need_add_list is inserted, time is:"+str(datetime.datetime.now()))
            #3.全部更新
            qtDept=is_department.objects.all()
            for dept in qtDept:
                if dept.department_id != 10000:
                    hr_dept_list=fetchall_to_dict("select DEPARTMENTKEY,DEPARTMENTNAME,MANAGER_ID from ir_dimdepartment where INUSE='Y' and DEPARTMENTKEY=%(DEPARTMENTKEY)s",params={"DEPARTMENTKEY":dept.department_id},db='hr')
                    if is_department.objects.filter(department_id=dept.department_id).exists():
                        is_department.objects.filter(department_id=dept.department_id).update(department_name=hr_dept_list[0]['DEPARTMENTNAME'],create_time=datetime.datetime.now())
            print('qtcompetence_is_department is compeletly updated,time is:'+str(datetime.datetime.now()))    
        # is_position
        #需要删除的数据
        need_delete_list=[]
        #需要增加的数据的列表
        need_add_list=[]
        #需要更新的数据
        need_update_list=[]
        qtPos=is_position.objects.all()
        if not qtPos:
            #首次插入为初始化插入    
            #初始化插入
            for i in range(0,len(pos_list)):
                #exec_sql('insert into iq.qtcompetence_is_position values (%(ID)s,%(JOB_CODE_ID)s,%(JOB_TITLE)s,%(TIME)s)',params={"ID":[i+1],"JOB_CODE_ID":pos_list[i]['JOB_CODE_ID'],"JOB_TITLE":pos_list[i]['JOB_TITLE'],"TIME":datetime.datetime.now()})
                is_position.objects.create(position_id=pos_list[i]['JOB_CODE_ID'],position_name=pos_list[i]['JOB_TITLE']) 
            #exec_sql('insert into iq.qtcompetence_is_position values (%(ID)s,%(JOB_CODE_ID)s,%(JOB_TITLE)s,%(TIME)s)',params={"ID":10000,"JOB_CODE_ID":10000,"JOB_TITLE":"NO POSITION","TIME":datetime.datetime.now()})     
            is_position.objects.create(id=10000,position_id=10000,position_name="NO POSITION")  
            print('qtcompetence_is_position is inserted at first time, time is :'+str(datetime.datetime.now()))
        else:
            #先获取目前最大索引id出去第10000条
            now_index=(is_position.objects.order_by('-id')[1:2])[0].id
            print(now_index)
            #1.遍历local对比远程端筛选出需要删除(本地有远程端没有)项进行删除
            for pos in qtPos:
                is_pos_exist=fetchall_to_dict("select JOB_CODE_ID,JOB_TITLE,INUSE from OEE.ETF_Dim_JobCode where INUSE='Y' and JOB_CODE_ID=%(JOB_CODE_ID)s",params={"JOB_CODE_ID":pos.position_id},db='hr')
                if not is_pos_exist:
                    #10000为手动添加No Department for django
                    if pos.position_id != 10000:
                        need_delete_list .append(pos.id)
                else:
                    if is_pos_exist[0]['INUSE'] == 'N':
                        need_delete_list.append(pos.id)
            print("is_position's need_delete_list is:"+str(need_delete_list))
            #删除本地多余数据USER_ID
            if need_delete_list:
                for id in need_delete_list:
                    #exec_sql('delete from iq.qtcompetence_is_position where position_id=%(JOB_CODE_ID)s',params={"JOB_CODE_ID":id})
                    is_position.objects.filter(id=id).delete()
                print("is_position's need_delete_list is deleted,time is:"+str(datetime.datetime.now()))
            #2.遍历远程数据库对比本地筛选出需要增加(远程端有本地没有)项进行增加
            for i in range(0,len(pos_list)):
                if not (is_position.objects.filter(position_id=pos_list[i]['JOB_CODE_ID']).exists() and pos_list[i]['INUSE'] == 'Y'):
                    need_add_list.append(pos_list[i]['JOB_CODE_ID'])
            print("is_position's need_add_list is:"+str(need_add_list))
            #直接插入本地
            if need_add_list:
                for i in range(0,len(need_add_list)):
                    hr_pos_list=fetchall_to_dict("select JOB_CODE_ID,JOB_TITLE from OEE.ETF_Dim_JobCode where INUSE='Y' and JOB_CODE_ID=%(JOB_CODE_ID)s",params={"JOB_CODE_ID":need_add_list[i]},db='hr')
                    print(hr_pos_list)
                    is_position.objects.create(id=now_index+i+1,position_id=hr_pos_list[0]['JOB_CODE_ID'],position_name=hr_pos_list[0]['JOB_TITLE'])
            print("is_position's need_add_list is inserted, time is :"+str(datetime.datetime.now()))
            #3.全部更新
            qtPos=is_position.objects.all()
            for pos in qtPos:
                if pos.position_id != 10000:
                    hr_pos_list=fetchall_to_dict("select JOB_CODE_ID,JOB_TITLE from OEE.ETF_Dim_JobCode where INUSE='Y' and JOB_CODE_ID=%(JOB_CODE_ID)s",params={"JOB_CODE_ID":pos.position_id},db='hr')
                    if is_position.objects.filter(position_id=hr_pos_list[0]['JOB_CODE_ID']).exists():
                        is_position.objects.filter(position_id=hr_pos_list[0]['JOB_CODE_ID']).update(position_name=hr_pos_list[0]['JOB_TITLE'],create_time=datetime.datetime.now())  
            print('is_position is completly updated,time is:'+str(datetime.datetime.now()))
                # is_user
        #is_user
        #需要删除的数据
        need_delete_list=[]
        #需要增加的数据的列表
        need_add_list=[]
        #需要更新的数据
        need_update_list=[]
        qtUser=is_user.objects.all()
        #初始化插入
        if not qtUser:
            #初始化插入
                for i in range(0,len(user_list)):
                    #email_id 需要处理一下
                    #职位与部门插入时候因为主键问题新建了连续主键列，所有对应主键外键关系需要转换一下
                    # is_department is_position 第一万条记录处理空数据手动创建
                    is_user.objects.create(password=user_list[i]['SYS_NAME'],is_superuser=0,user_id=user_list[i]['ID'],user_name=user_list[i]['BADGE_NO'],employ_name=user_list[i]['USER_NAME'],
                                            email=normallizeEmail(user_list[i]['EMAIL_ID']),is_active=1,is_admin=0,department_id=deptID_to_ID(user_list[i]['DEPARTMENTKEY']),position_id=posID_to_ID(user_list[i]['JOB_CODE_ID']))
                is_user.objects.create(id=100000,password="NO NAME",is_superuser=0,user_id=100000,user_name="NO BADGE",employ_name="NO USER NAME",
                                        email="NO EMAIL",is_active=1,is_admin=0,department_id=10000,position_id=10000)
                print('is_user is inserted at first time,time is:'+str(datetime.datetime.now()))
        else:
            #先获取目前最大索引id
            now_index=(is_user.objects.order_by('-id')[1:2])[0].id
            print(now_index)
            #1.遍历local对比远程端筛选出需要删除(本地有远程端没有)项进行删除
            for user in qtUser:
                is_user_exist=fetchall_to_dict("select ID,INUSE from ir_dimUser where id=%(ID)s and INUSE='Y'",params={"ID":user.user_id},db='hr')
                if not is_user_exist:
                    #100000与100001为手动添加No Name No Badge for django
                    if user.user_id != 100000:
                        need_delete_list.append(user.id)
                else:
                    if is_user_exist[0]['INUSE'] == 'N':
                        need_delete_list.append(user.id)
            print("is_user's need_delete_list are:"+str(need_delete_list))
            #删除本地多余数据USER_ID
            if need_delete_list:
                for id in need_delete_list:
                    is_user.objects.filter(id=id).delete()
                    #删除is_competition_score中死数据
                    is_competition_score.objects.filter(user_id=id).delete()
                    #exec_sql('delete from iq.qtcompetence_is_user where user_id=%(USER_ID)s',params={"USER_ID":id})
            print("is_user's need_delete_list is deleted,time is:"+str(datetime.datetime.now()))
            print("is_competition_score need_delete_list is deleted,time is:"+str(datetime.datetime.now()))
            #2.遍历远程数据库对比本地筛选出需要增加(远程端有本地没有)项进行增加
            for i in range(0,len(user_list)):
                if not (is_user.objects.filter(user_id=user_list[i]['ID']).exists() and user_list[i]['INUSE'] == 'Y'):
                    need_add_list.append(user_list[i]['ID'])
            print("is_user's and need_add_list are:"+str(need_add_list))
            #直接插入本地
            if need_add_list:
                for i in range(0,len(need_add_list)):
                    hr_user_list=fetchall_to_dict("select ID,SYS_NAME,BADGE_NO,USER_NAME,EMAIL_ID,DEPARTMENTKEY,JOB_CODE_ID,LEVEL_1_R,LEVEL_2_R,LEVEL_3_R,LEVEL_4_R from ir_dimUser where INUSE='Y' and ID=%(ID)s",params={"ID":need_add_list[i]},db='hr')
                    print(hr_user_list)
                    is_user.objects.create(id=now_index+i+1,password=hr_user_list[0]['SYS_NAME'],is_superuser=0,user_id=hr_user_list[0]['ID'],user_name=hr_user_list[0]['BADGE_NO'],employ_name=hr_user_list[0]['USER_NAME'],
                                            email=normallizeEmail(hr_user_list[0]['EMAIL_ID']),is_active=1,is_admin=0,department_id=deptID_to_ID(hr_user_list[0]['DEPARTMENTKEY']),position_id=posID_to_ID(hr_user_list[0]['JOB_CODE_ID']))
                print("is_user's need_add_list are inserted, time is :"+str(datetime.datetime.now()))
            #3.全部更新
            qtUser=is_user.objects.all()
            for user in qtUser:
                if user.user_id != 100000:
                    hr_user_list=fetchall_to_dict("select ID,SYS_NAME,BADGE_NO,USER_NAME,EMAIL_ID,DEPARTMENTKEY,JOB_CODE_ID,LEVEL_1_R,LEVEL_2_R,LEVEL_3_R,LEVEL_4_R from ir_dimUser where INUSE='Y' and ID=%(ID)s",params={"ID":user.user_id},db='hr')
                    if is_user.objects.filter(user_id=user.user_id).exists():
                        is_user.objects.filter(user_id=user.user_id).update(password=hr_user_list[0]['SYS_NAME'],user_id=hr_user_list[0]['ID'],user_name=hr_user_list[0]['BADGE_NO'],employ_name=hr_user_list[0]['USER_NAME'],
                                                email=normallizeEmail(hr_user_list[0]['EMAIL_ID']),department_id=deptID_to_ID(hr_user_list[0]['DEPARTMENTKEY']),position_id=posID_to_ID(hr_user_list[0]['JOB_CODE_ID']),create_time=datetime.datetime.now())
                        #此刻应该同时更新is_与competence有user_id的都需要更新
            print("is_user is completly updated,time is:"+str(datetime.datetime.now()))
        # r_level
        #需要删除的数据
        need_delete_list=[]
        #需要增加的数据的列表
        need_add_list=[]
        #需要更新的数据
        need_update_list=[]
        qtLevel=r_level.objects.all()
        #初始化插入
        if not qtLevel:
            #初始化插入
            for i in range(0,len(user_list)):
                #exec_sql('insert into iq.qtcompetence_r_level(level_id,create_time,L1_id,L2_id,L3_id,L4_id,user_id) values (%(ID)s,%(TIME)s,%(LEVEL_1_R)s,%(LEVEL_2_R)s,%(LEVEL_3_R)s,%(LEVEL_4_R)s,%(USER_ID)s)',
                # params={"ID":i+1,"USER_ID":(user_list[i]['ID']),"LEVEL_1_R":userID_to_ID(user_list[i]['LEVEL_1_R']),"LEVEL_2_R":userID_to_ID(user_list[i]['LEVEL_2_R']),
                # "LEVEL_3_R":userID_to_ID(user_list[i]['LEVEL_3_R']),"LEVEL_4_R":userID_to_ID(user_list[i]['LEVEL_4_R']),"TIME":datetime.datetime.now()})
                r_level.objects.create(L1_id=userID_to_ID(user_list[i]['LEVEL_1_R']),L2_id=userID_to_ID(user_list[i]['LEVEL_2_R']),L3_id=userID_to_ID(user_list[i]['LEVEL_3_R']),
                                        L4_id=userID_to_ID(user_list[i]['LEVEL_4_R']),user_id=userID_to_ID(user_list[i]['ID']))
            # exec_sql('insert into iq.qtcompetence_is_user values (%(USER_ID)s,%(SYS_NAME)s,Null,0,%(ID)s,%(BADGE_NO)s,%(USER_NAME)s,%(EMAIL_ID)s,1,0,%(TIME)s,%(DEPARTMENTKEY)s,%(JOB_CODE_ID)s)',
            #     params={"USER_ID":100000,"ID":100000,"SYS_NAME":'NO NAME',"BADGE_NO":'NO BADGE',"USER_NAME":'NO USER_NAME',
            #     "EMAIL_ID":'NO EMAIL_ID',"DEPARTMENTKEY":10000,"JOB_CODE_ID":10000,"TIME":datetime.datetime.now()})
            print('r_level is inserted at first time,time is:'+str(datetime.datetime.now()))
        else:
            #先获取目前最大索引id
            now_index=(r_level.objects.order_by('-level_id')[0:1])[0].level_id
            print(now_index)
            #1.遍历local对比远程端筛选出需要删除(本地有远程端没有)项进行删除
            for level in qtLevel:
                is_level_exist=fetchall_to_dict("select ID, INUSE from ir_dimUser where id=%(ID)s and INUSE='Y'",params={"ID":ID_to_userID(level.user_id)},db='hr')
                if not is_level_exist:
                    need_delete_list.append(level.user_id)
                else:
                    if is_level_exist[0]['INUSE'] == 'N':
                        need_delete_list.append(level.user_id)
            print("r_level's need_delete_list are:"+str(need_delete_list))
            #删除本地多余数据USER_ID
            if need_delete_list:
                for id in need_delete_list:
                    #r_level中插入的的经过转换的user_id与服务器端ID不同
                    r_level.objects.filter(user_id=id).delete()
                    #exec_sql('delete from iq.qtcompetence_r_level where user_id=%(USER_ID)s',params={"USER_ID":id})
                print("r_level's need_delete_list is deleted,time is:"+str(datetime.datetime.now()))
            #2.遍历远程数据库对比本地筛选出需要增加(远程端有本地没有)项进行增加
            #r_level
            for i in range(0,len(user_list)):
                if not (r_level.objects.filter(user_id=userID_to_ID(user_list[i]['ID'])).exists() and user_list[i]['INUSE'] == 'Y'):
                    need_add_list.append(user_list[i]['ID'])
            print("r_level's need_add_list are:"+str(need_add_list))
            #直接插入本地
            if need_add_list:
                for i in range(0,len(need_add_list)):
                    hr_user_list=fetchall_to_dict("select ID,SYS_NAME,BADGE_NO,USER_NAME,EMAIL_ID,DEPARTMENTKEY,JOB_CODE_ID,LEVEL_1_R,LEVEL_2_R,LEVEL_3_R,LEVEL_4_R from ir_dimUser where INUSE='Y' and ID=%(ID)s",params={"ID":need_add_list[i]},db='hr')
                    print(hr_user_list)
                    r_level.objects.create(level_id=now_index+i+1,L1_id=userID_to_ID(hr_user_list[0]['LEVEL_1_R']),L2_id=userID_to_ID(hr_user_list[0]['LEVEL_2_R']),L3_id=userID_to_ID(hr_user_list[0]['LEVEL_3_R']),
                                            L4_id=userID_to_ID(hr_user_list[0]['LEVEL_4_R']),user_id=userID_to_ID(hr_user_list[0]['ID']))
                    # exec_sql('insert into iq.qtcompetence_is_user values (%(USER_ID)s,%(SYS_NAME)s,Null,0,%(ID)s,%(BADGE_NO)s,%(USER_NAME)s,%(EMAIL_ID)s,1,0,%(TIME)s,%(DEPARTMENTKEY)s,%(JOB_CODE_ID)s)',
                    #     params={"USER_ID":now_index+i+1,"ID":need_add_list[i],"SYS_NAME":hr_user_list[0]['SYS_NAME'],"BADGE_NO":hr_user_list[0]['BADGE_NO'],"USER_NAME":hr_user_list[0]['USER_NAME'],
                    #     "EMAIL_ID":normallizeEmail(hr_user_list[0]['EMAIL_ID']),"DEPARTMENTKEY":deptID_to_ID(hr_user_list[0]['DEPARTMENTKEY']),"JOB_CODE_ID":posID_to_ID(hr_user_list[0]['JOB_CODE_ID']),"TIME":datetime.datetime.now()})
                    # exec_sql('insert into iq.qtcompetence_r_level(level_id,create_time,L1_id,L2_id,L3_id,L4_id,user_id) values (%(ID)s,%(TIME)s,%(LEVEL_1_R)s,%(LEVEL_2_R)s,%(LEVEL_3_R)s,%(LEVEL_4_R)s,%(USER_ID)s)',
                    # params={"ID":now_index+i+1,"USER_ID":(hr_user_list[0]['ID']),"LEVEL_1_R":userID_to_ID(hr_user_list[0]['LEVEL_1_R']),"LEVEL_2_R":userID_to_ID(hr_user_list[0]['LEVEL_2_R']),
                    # "LEVEL_3_R":userID_to_ID(hr_user_list[0]['LEVEL_3_R']),"LEVEL_4_R":userID_to_ID(hr_user_list[0]['LEVEL_4_R']),"TIME":datetime.datetime.now()})
                print("r_level's need_add_list are inserted, time is :"+str(datetime.datetime.now()))
            #3.全部更新
            qtLevel=r_level.objects.all()
            for level in qtLevel:
                hr_user_list=fetchall_to_dict("select ID,SYS_NAME,BADGE_NO,USER_NAME,EMAIL_ID,DEPARTMENTKEY,JOB_CODE_ID,LEVEL_1_R,LEVEL_2_R,LEVEL_3_R,LEVEL_4_R from ir_dimUser where INUSE='Y' and ID=%(ID)s",params={"ID":ID_to_userID(level.user_id)},db='hr')
                if  r_level.objects.filter(user_id=level.user_id).exists():
                    r_level.objects.filter(user_id=level.user_id).update(L1_id=userID_to_ID(hr_user_list[0]['LEVEL_1_R']),L2_id=userID_to_ID(hr_user_list[0]['LEVEL_2_R']),L3_id=userID_to_ID(hr_user_list[0]['LEVEL_3_R']),
                                            L4_id=userID_to_ID(hr_user_list[0]['LEVEL_4_R']),user_id=userID_to_ID(hr_user_list[0]['ID']),create_time=datetime.datetime.now())    
                    # exec_sql('update iq.qtcompetence_is_user set user_name=%(BADGE_NO)s,password=%(SYS_NAME)s,employ_name =%(USER_NAME)s,email =%(EMAIL_ID)s,department_id=%(DEPARTMENTKEY)s,position_id =%(JOB_CODE_ID)s,create_time=%(TIME)s where user_id=%(USER_ID)s',
                    #     params={"USER_ID":user.user_id,"SYS_NAME":hr_user_list[0]['SYS_NAME'],"BADGE_NO":hr_user_list[0]['BADGE_NO'],"USER_NAME":hr_user_list[0]['USER_NAME'],
                    #     "EMAIL_ID":normallizeEmail(hr_user_list[0]['EMAIL_ID']),"DEPARTMENTKEY":deptID_to_ID(hr_user_list[0]['DEPARTMENTKEY']),"JOB_CODE_ID":posID_to_ID(hr_user_list[0]['JOB_CODE_ID']),"TIME":datetime.datetime.now()})
            print("r_level is completly updated,time is:"+str(datetime.datetime.now()))
    register_events(scheduler)
    scheduler.start()
except Exception as e:
    print(e)
    # 有错误就停止定时器
    scheduler.shutdown()

# some templates about pages using
# class register_user(FormView):
#     pass
#     template_name = 'QtCompetence/register_user.html'
#     form_class = register_form
#     success_url = reverse_lazy('QtCompetence:home')

# def register_user(request):
    # if request.method == 'POST':
    #     # 实例化form对象时候，把post提交过来的数据直接传进去
    #     regForm = RegisterUserForm(request.POST)
    #     # 调用regForm校验数据的方法
    #     if regForm.is_valid():
    #         userName = regForm.cleaned_data['userName']
    #         email = regForm.cleaned_data['email']
    #         phone = regForm.cleaned_data['phone']
    #         passWord = regForm.cleaned_data['passWord']
    #         encryptPwd = make_password(passWord)
    #         is_user.objects.create(userName=userName, email=email, phone=phone,
    #                                 passWord=encryptPwd)
    #         return render(request, 'QtCompetence/register_user.html', {"result": "恭喜您注册成功",
    #                                                                 "formDisplay": json.dumps({'display': '0'})})
    # else:
    #     regForm = RegisterUserForm()
    # return render(request, QtCompetence/register_user.html', {"regForm": regForm,
    #                                                             "formDisplay": json.dumps({'display': '1'})})